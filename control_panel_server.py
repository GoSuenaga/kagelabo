"""
VANTAN Video Studio — コントロールパネル サーバー v1.2
XLSX設計図カートリッジ版（RAG Creative Studio パターン）

使い方: python3 control_panel_server.py
→ http://localhost:8888 をブラウザで開く
"""
import os, json, glob, time, random, threading, re, shutil
from datetime import datetime
from fastapi import FastAPI, BackgroundTasks, UploadFile, File
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from starlette.responses import Response
from openpyxl import load_workbook
import uvicorn

ROOT = os.path.dirname(os.path.abspath(__file__))
VANTAN = os.path.join(ROOT, "apps", "vantan-video")
CLIENTS = os.path.join(ROOT, "clients")  # ロゴ・SE・BGM はルート直下の clients/
BRIEFS_DIR = os.path.join(VANTAN, "briefs")
STATE_DIR = os.path.join(VANTAN, "state")
META_PATH = os.path.join(VANTAN, "cartridge_meta.json")
os.makedirs(BRIEFS_DIR, exist_ok=True)
os.makedirs(STATE_DIR, exist_ok=True)

# ============================================================
# カートリッジシステム（XLSX ベース — RAG パターン）
# ============================================================
REGISTRY_PATH = os.path.join(STATE_DIR, "cartridge_registry.json")

def load_registry():
    if os.path.exists(REGISTRY_PATH):
        with open(REGISTRY_PATH, encoding="utf-8") as f:
            return json.load(f)
    return {"next_id": 1, "cartridges": {}}

def save_registry(reg):
    with open(REGISTRY_PATH, "w", encoding="utf-8") as f:
        json.dump(reg, f, ensure_ascii=False, indent=2)

def register_cartridge(filename):
    """XLSX ファイルを登録して CID を返す（同名なら既存IDを返す）"""
    reg = load_registry()
    for cid, info in reg.get("cartridges", {}).items():
        if info["filename"] == filename:
            return cid
    num = reg.get("next_id", 1)
    cid = f"C{num:03d}"
    reg.setdefault("cartridges", {})[cid] = {
        "filename": filename,
        "registered_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    reg["next_id"] = num + 1
    save_registry(reg)
    return cid

def load_meta():
    if os.path.exists(META_PATH):
        with open(META_PATH, encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_meta(meta):
    with open(META_PATH, "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)

def get_active_cid():
    return load_meta().get("cartridge_id", "")

def set_active_cid(cid, filename="", pattern_count=0):
    save_meta({
        "cartridge_id": cid,
        "loaded_file": filename,
        "loaded_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "pattern_count": pattern_count,
    })

def _patterns_path(cid):
    return os.path.join(STATE_DIR, f"{cid}_patterns.json")

def _state_path(cid):
    return os.path.join(STATE_DIR, f"{cid}_state.json")

def _corrections_path(cid):
    return os.path.join(STATE_DIR, f"{cid}_corrections.json")

def load_cartridge_patterns(cid):
    p = _patterns_path(cid)
    if os.path.exists(p):
        with open(p, encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_cartridge_patterns(cid, patterns):
    with open(_patterns_path(cid), "w", encoding="utf-8") as f:
        json.dump(patterns, f, ensure_ascii=False, indent=2)

def cartridge_output_base(cid):
    """カートリッジごとの出力ディレクトリ"""
    d = os.path.join(VANTAN, "output", cid)
    os.makedirs(d, exist_ok=True)
    return d

# ============================================================
# XLSX パーサー（VANTAN 台本カラム構成）
# ============================================================
COLUMN_MAP = {
    "No": "no",
    "スクール名称": "school",
    "分野": "field",
    "子ども": "child",
    "カット#": "num",
    "カットタイプ": "cut_type",
    "ナレーション": "narration",
    "テロップ": "telop",
    "ロゴ表示": "logo",
    "ロゴファイルパス": "logo_path",
    "映像プロンプト（日本語）": "prompt_jp",
    "映像プロンプト（EN）": "prompt_en",
    "LP": "lp",
    "動画のムード": "mood",
}

def parse_xlsx_to_patterns(filepath):
    """XLSX を読み込んでパターン辞書に変換"""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    # シート名検索: 「台本」「brief」を優先、なければ先頭
    ws = None
    for name in wb.sheetnames:
        if "台本" in name or "brief" in name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return {}

    # ヘッダーマッピング
    headers = [str(h).strip() if h else "" for h in rows[0]]
    col_idx = {}
    for i, h in enumerate(headers):
        if h in COLUMN_MAP:
            col_idx[COLUMN_MAP[h]] = i
        else:
            # fuzzy: ヘッダーの部分一致
            for pattern, key in COLUMN_MAP.items():
                if pattern in h:
                    col_idx[key] = i
                    break

    def cell(row, key):
        idx = col_idx.get(key)
        if idx is None or idx >= len(row):
            return ""
        return str(row[idx]).strip() if row[idx] is not None else ""

    # パターン構築
    patterns = {}
    current_no, current_school, current_field, current_child = "", "", "", ""
    for row in rows[1:]:
        no_val = cell(row, "no")
        if no_val and no_val != "None":
            current_no = no_val
            current_school = cell(row, "school") or current_school
            current_field = cell(row, "field") or current_field
            current_child = cell(row, "child") or current_child

        cut_num = cell(row, "num")
        if not cut_num or cut_num == "None":
            continue

        key = f"no{current_no.zfill(2)}"
        if key not in patterns:
            patterns[key] = {
                "school": current_school,
                "field": current_field,
                "child": current_child,
                "cuts": [],
            }
        patterns[key]["cuts"].append({
            "num": cut_num,
            "cut_type": cell(row, "cut_type"),
            "narration": cell(row, "narration"),
            "telop": cell(row, "telop"),
            "logo": cell(row, "logo"),
            "logo_path": cell(row, "logo_path"),
            "prompt_jp": cell(row, "prompt_jp"),
            "prompt_en": cell(row, "prompt_en"),
            "lp": cell(row, "lp"),
            "mood": cell(row, "mood"),
        })

    return patterns

# ============================================================
# 状態管理（カートリッジ対応）
# ============================================================
_state_lock = threading.Lock()

def load_state():
    cid = get_active_cid()
    if not cid:
        return {}
    sf = _state_path(cid)
    with _state_lock:
        try:
            if os.path.exists(sf):
                with open(sf, encoding="utf-8") as f:
                    return json.load(f)
        except (json.JSONDecodeError, IOError):
            pass
        return {}

def save_state(state):
    cid = get_active_cid()
    if not cid:
        return
    sf = _state_path(cid)
    with _state_lock:
        tmp = sf + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(state, f, ensure_ascii=False, indent=2)
        os.replace(tmp, sf)

def get_pattern_state(pat_key):
    state = load_state()
    return state.get(pat_key, {"status": "idle", "progress": "", "message": ""})

def set_pattern_state(pat_key, **kwargs):
    state = load_state()
    if pat_key not in state:
        state[pat_key] = {"status": "idle", "progress": "", "message": "", "stop_requested": False}
    state[pat_key].update(kwargs)
    state[pat_key]["updated"] = datetime.now().strftime("%H:%M:%S")
    save_state(state)

def get_cut_state(pat_key, cut_num):
    state = load_state()
    pat = state.get(pat_key, {})
    cuts = pat.get("cuts_qc", {})
    return cuts.get(str(cut_num), {
        "qc_status": "pending",
        "narration_override": None,
        "telop_override": None,
        "prompt_jp_override": None,
        "prompt_en_override": None,
    })

def set_cut_state(pat_key, cut_num, **kwargs):
    state = load_state()
    if pat_key not in state:
        state[pat_key] = {"status": "idle", "progress": "", "message": "", "stop_requested": False}
    if "cuts_qc" not in state[pat_key]:
        state[pat_key]["cuts_qc"] = {}
    cn = str(cut_num)
    if cn not in state[pat_key]["cuts_qc"]:
        state[pat_key]["cuts_qc"][cn] = {"qc_status": "pending"}
    state[pat_key]["cuts_qc"][cn].update(kwargs)
    save_state(state)

# ============================================================
# アクティブパターン（カートリッジから読み込み）
# ============================================================
PATTERNS = {}

def reload_active_patterns():
    """アクティブカートリッジのパターンをメモリに読み込む"""
    global PATTERNS
    cid = get_active_cid()
    if cid:
        PATTERNS = load_cartridge_patterns(cid)
    else:
        PATTERNS = {}
    return PATTERNS

def get_output_base():
    """現在のカートリッジの出力ディレクトリ"""
    cid = get_active_cid()
    if cid:
        return cartridge_output_base(cid)
    return os.path.join(VANTAN, "output", "_no_cartridge")

# ============================================================
# 起動時: アクティブカートリッジを読み込む
# ============================================================
cid = get_active_cid()
if cid:
    print(f"カートリッジ {cid} を読み込み中...")
    PATTERNS = load_cartridge_patterns(cid)
    meta = load_meta()
    print(f"  ✓ {meta.get('loaded_file', '?')} ({len(PATTERNS)} パターン)")
else:
    print("アクティブカートリッジなし（UIから XLSX をインポートしてください）")

# ============================================================
# 出力チェック
# ============================================================
def check_outputs(pat_key):
    base = os.path.join(get_output_base(), pat_key)
    vid_dir, aud_dir = os.path.join(base, "videos"), os.path.join(base, "audio")
    videos = sorted([f for f in os.listdir(vid_dir) if f.endswith('.mp4')]) if os.path.isdir(vid_dir) else []
    audio = sorted([f for f in os.listdir(aud_dir) if f.endswith('.mp3')]) if os.path.isdir(aud_dir) else []
    finals = sorted(glob.glob(os.path.join(base, "final*.mp4")))
    return {"videos": videos, "audio": audio, "final": finals[-1] if finals else None}

def count_cut_versions(pat_key, cut_num):
    base = os.path.join(get_output_base(), pat_key, "videos")
    if not os.path.isdir(base):
        return 0
    prefix = f"カット{str(cut_num).zfill(2)}"
    return len([f for f in os.listdir(base) if f.startswith(prefix) and f.endswith('.mp4')])

def list_cut_versions(pat_key, cut_num):
    base = os.path.join(get_output_base(), pat_key, "videos")
    if not os.path.isdir(base):
        return []
    prefix = f"カット{str(cut_num).zfill(2)}"
    files = sorted([f for f in os.listdir(base) if f.startswith(prefix) and f.endswith('.mp4')])
    cid = get_active_cid()
    return [{"name": f, "url": f"/media/{cid}/{pat_key}/videos/{f}"} for f in files]

# ============================================================
# 修正ログ
# ============================================================
def log_correction(pat_key, cut_num, field, original, corrected, reason=""):
    cid = get_active_cid()
    if not cid:
        return
    path = _corrections_path(cid)
    entries = []
    if os.path.exists(path):
        with open(path, encoding="utf-8") as f:
            entries = json.load(f)
    entries.append({
        "pattern": pat_key, "cut": str(cut_num), "field": field,
        "original": original, "corrected": corrected, "reason": reason,
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    })
    with open(path, "w", encoding="utf-8") as f:
        json.dump(entries, f, ensure_ascii=False, indent=2)

# ============================================================
# ロゴマップ
# ============================================================
LOGO_MAP = {
    "バンタンデザイン研究所": "clients/vantan/2026/VDI/専門部/2025_VDI_PRO_logo_10801080.jpg",
    "バンタンゲームアカデミー": "clients/vantan/2026/VGA/専門部/2025_VGA_PRO_logo_10801080.jpg",
    "ヴィーナスアカデミー": "clients/vantan/2026/VA/専門部/2025_VA_PRO_logo_10801080.jpg",
    "バンタンクリエイターアカデミー": "clients/vantan/2026/VCA/専門部/2026_VCA_PRO_logo_10801080.jpg.jpg",
    "レコールバンタン": "clients/vantan/2026/LV/専門部/2023_LV_PRO_logo_10801080.jpg",
    "KADOKAWAアニメ・声優アカデミー": "clients/vantan/2026/KAA/専門部/2026_KAA_PRO_logo_10801080.jpg",
    "KADOKAWAマンガアカデミー": "clients/vantan/2026/KMA/専門部/2026_KMA_PRO_logo_10801080.jpg",
    "バンタンミュージックアカデミー": "clients/vantan/2026/VMA/専門部/2026_VMA_PRO_logo_10801080.jpg",
}

CINEMA_SUFFIX = (
    "Shot on a high-end cinema camera with anamorphic lens. "
    "Shallow depth of field, natural bokeh. "
    "Documentary film aesthetic, warm cinematic color grading. "
    "Subtle natural camera movement. No text, no logos, no signage."
)

# ============================================================
# 生成ロジック（バックグラウンド）— カートリッジ対応
# ============================================================
def run_pipeline(pat_key, steps=None):
    from dotenv import load_dotenv
    import requests as req
    load_dotenv(os.path.join(ROOT, ".env"))
    fal_key = os.getenv("FAL_API_KEY")
    creatomate_key = os.getenv("CREATOMATE_API_KEY")

    if steps is None:
        steps = ["video", "narration", "compose"]

    pat = PATTERNS.get(pat_key)
    if not pat:
        set_pattern_state(pat_key, status="error", message="パターンが見つかりません")
        return

    cuts = pat["cuts"]
    out_dir = os.path.join(get_output_base(), pat_key)
    os.makedirs(os.path.join(out_dir, "videos"), exist_ok=True)
    os.makedirs(os.path.join(out_dir, "audio"), exist_ok=True)

    set_pattern_state(pat_key, status="running", progress="0/?", message="開始中...", stop_requested=False)

    def should_stop():
        s = load_state()
        return s.get(pat_key, {}).get("stop_requested", False)

    def upload(filepath, content_type):
        full = os.path.join(VANTAN, filepath) if not os.path.isabs(filepath) else filepath
        with open(full, 'rb') as f:
            d = f.read()
        try:
            init = req.post(
                "https://rest.alpha.fal.ai/storage/upload/initiate",
                headers={"Authorization": f"Key {fal_key}", "Content-Type": "application/json"},
                json={"file_name": os.path.basename(filepath), "content_type": content_type},
                timeout=30,
            )
            if init.status_code == 200:
                req.put(init.json()["upload_url"], data=d, headers={"Content-Type": content_type}, timeout=60)
                return init.json()["file_url"]
        except Exception:
            pass
        try:
            r = req.post("https://0x0.st", files={"file": (os.path.basename(filepath), d, content_type)}, timeout=60)
            if r.status_code == 200:
                return r.text.strip()
        except Exception:
            pass
        raise Exception(f"ファイルアップロード失敗: {os.path.basename(filepath)}")

    try:
        # ===== Step 1: 動画生成 =====
        # fal.ai ファースト → Gemini フォールバック
        if "video" in steps:
            from google import genai
            import fal_client

            # --- API キー ---
            _fal_key = os.getenv("FAL_API_KEY", "")
            _gemini_keys = []
            for _ki in range(1, 10):
                _k = os.getenv(f"GEMINI_API_KEY_{_ki}", "")
                if _k:
                    _gemini_keys.append(_k)

            # --- モデル設定 ---
            # パターンまたはグローバルの video_model 設定を取得
            _video_model = pat.get("video_model", "fal-ai/veo3.1")  # デフォルト: fal Veo 3.1

            # --- プロバイダーリスト構築 ---
            _providers = []
            if _video_model == "gemini-veo3.1":
                # Gemini 優先モード
                if len(_gemini_keys) >= 2:
                    _providers.append(("gemini", 1, "veo-3.1-generate-preview"))
                if len(_gemini_keys) >= 3:
                    _providers.append(("gemini", 2, "veo-3.1-generate-preview"))
                if len(_gemini_keys) >= 1:
                    _providers.append(("gemini", 0, "veo-3.1-generate-preview"))
                if _fal_key:
                    _providers.append(("fal", "fal-ai/veo3.1"))  # フォールバック
            else:
                # fal.ai ファースト（デフォルト）
                if _fal_key:
                    _providers.append(("fal", _video_model))
                # Gemini フォールバック
                if len(_gemini_keys) >= 2:
                    _providers.append(("gemini", 1, "veo-3.1-generate-preview"))
                if len(_gemini_keys) >= 3:
                    _providers.append(("gemini", 2, "veo-3.1-generate-preview"))
                if len(_gemini_keys) >= 1:
                    _providers.append(("gemini", 0, "veo-3.1-generate-preview"))
            if not _providers:
                set_pattern_state(pat_key, status="error", message="FAL_API_KEY も GEMINI_API_KEY もありません")
                return

            _provider_idx = 0

            def _provider_label(idx=None):
                if idx is None:
                    idx = _provider_idx
                p = _providers[idx]
                if p[0] == "fal":
                    return p[1]  # e.g. "fal-ai/veo3.1"
                return f"Gemini KEY_{p[1]+1} ({p[2]})"

            def _rotate_provider(err_msg):
                nonlocal _provider_idx
                _provider_idx += 1
                if _provider_idx >= len(_providers):
                    return False
                set_pattern_state(pat_key, message=f"→ {_provider_label()} に切替（{err_msg[:50]}）")
                return True

            def _is_retryable(err_msg):
                lower = err_msg.lower()
                return ("quota" in lower or "429" in lower or "resource_exhausted" in lower
                        or "rate" in lower or "動画データなし" in err_msg or "動画ファイルが空" in err_msg)

            # ── fal.ai 生成 ──
            def _generate_video_fal(prompt, vid_path, cut_num, model_id):
                """fal.ai 経由で動画生成"""
                _tag = f"[{model_id}]"
                os.environ["FAL_KEY"] = _fal_key
                set_pattern_state(pat_key, message=f"カット{cut_num} {_tag} 生成中...")

                fal_args = {"prompt": prompt, "aspect_ratio": "9:16", "duration": "8s"}
                # image-to-video の場合は別途画像URLが必要（将来対応）

                result = fal_client.subscribe(model_id, arguments=fal_args, with_logs=False)
                video_url = result.get("video", {}).get("url", "")
                if not video_url:
                    raise RuntimeError(f"fal.ai: 動画URLが返りませんでした")
                vid_data = req.get(video_url, timeout=180)
                vid_data.raise_for_status()
                with open(vid_path, 'wb') as f:
                    f.write(vid_data.content)
                vid_size = len(vid_data.content)
                if vid_size == 0:
                    os.remove(vid_path)
                    raise RuntimeError("動画ファイルが空（0KB）")
                set_cut_state(pat_key, cut_num, video_model=model_id)
                set_pattern_state(pat_key, message=f"カット{cut_num} ✓ {_tag} ({vid_size//1024}KB)")

            # ── Gemini API 生成 ──
            def _generate_video_gemini(prompt, vid_path, cut_num, key_idx, model_name):
                """Gemini API 経由で動画生成"""
                _tag = f"[{model_name} / Gemini KEY_{key_idx+1}]"
                gclient = genai.Client(api_key=_gemini_keys[key_idx])
                set_pattern_state(pat_key, message=f"カット{cut_num} {_tag} 生成中...")
                op = gclient.models.generate_videos(
                    model=model_name,
                    prompt=prompt,
                    config=genai.types.GenerateVideosConfig(aspect_ratio="9:16", number_of_videos=1),
                )
                for poll_i in range(90):
                    if should_stop():
                        raise InterruptedError("停止リクエスト")
                    time.sleep(10)
                    op = gclient.operations.get(op)
                    set_pattern_state(pat_key, message=f"カット{cut_num} {_tag} {'完了' if op.done else '生成中'} ({poll_i+1})")
                    if op.done:
                        if op.response and op.response.generated_videos:
                            video = op.response.generated_videos[0]
                            vid_data = gclient.files.download(file=video.video)
                            with open(vid_path, 'wb') as f:
                                for chunk in vid_data:
                                    if isinstance(chunk, (bytes, bytearray)):
                                        f.write(chunk)
                                    elif isinstance(chunk, int):
                                        f.write(bytes([chunk]))
                                    else:
                                        f.write(bytes(chunk))
                            vid_size = os.path.getsize(vid_path)
                            if vid_size > 0:
                                set_cut_state(pat_key, cut_num, video_model=f"{model_name} / Gemini KEY_{key_idx+1}")
                                set_pattern_state(pat_key, message=f"カット{cut_num} ✓ {_tag} ({vid_size//1024}KB)")
                                return
                            else:
                                os.remove(vid_path)
                                raise RuntimeError("動画ファイルが空（0KB）")
                        else:
                            raise RuntimeError("動画データなし（ポリシー拒否の可能性）")
                raise TimeoutError("タイムアウト（15分）")

            # ── カットループ ──
            set_pattern_state(pat_key, message=f"Step 1/3: 動画生成（{_provider_label(0)} 優先）")
            for i, cut in enumerate(cuts):
                if should_stop():
                    set_pattern_state(pat_key, status="stopped", message="停止")
                    return
                num = cut["num"]
                vid_path = os.path.join(out_dir, "videos", f"カット{num.zfill(2)}.mp4")
                if os.path.exists(vid_path):
                    set_pattern_state(pat_key, progress=f"動画 {i+1}/{len(cuts)}", message=f"カット{num} スキップ（既存）")
                    continue

                cut_st = get_cut_state(pat_key, num)
                prompt_en = cut_st.get("prompt_en_override") or cut.get("prompt_en", "")
                if not prompt_en:
                    continue
                prompt_en = re.sub(r'(?i)(the camera work is slightly shaky.*?\.)', '', prompt_en)
                prompt_en = re.sub(r'(?i)(the shot looks raw.*?\.)', '', prompt_en)
                prompt_en = re.sub(r'(?i)(natural lighting, authentic social media aesthetic\.?)', '', prompt_en)
                prompt_en = re.sub(r'(?i)(as if shot on an? iphone.*?\.)', '', prompt_en)
                full_prompt = f"{prompt_en.strip()} {CINEMA_SUFFIX}"

                set_pattern_state(pat_key, progress=f"動画 {i+1}/{len(cuts)}", message=f"カット{num} {_provider_label()} で生成...")

                # 全プロバイダーを試行、同じカットをリトライ
                _saved_provider_idx = _provider_idx
                cut_success = False
                while _provider_idx < len(_providers):
                    p = _providers[_provider_idx]
                    try:
                        if p[0] == "fal":
                            _generate_video_fal(full_prompt, vid_path, num, p[1])
                        else:
                            _generate_video_gemini(full_prompt, vid_path, num, p[1], p[2])
                        cut_success = True
                        break
                    except InterruptedError:
                        return
                    except Exception as e:
                        err_msg = str(e)[:200]
                        if _is_retryable(err_msg):
                            if not _rotate_provider(err_msg):
                                set_pattern_state(pat_key, message=f"カット{num} ⚠ 全プロバイダー失敗、スキップ")
                                _provider_idx = _saved_provider_idx
                                break
                            continue
                        else:
                            set_pattern_state(pat_key, status="error", message=f"カット{num} エラー: {err_msg}")
                            return

                if not cut_success:
                    set_pattern_state(pat_key, message=f"カット{num} スキップ、次へ...")

        # ===== Step 2: ナレーション生成（ElevenLabs） =====
        if "narration" in steps:
            set_pattern_state(pat_key, message="Step 2/3: ナレーション生成（ElevenLabs）")
            elevenlabs_key = os.getenv("ELEVENLABS_API_KEY")
            voice_id = "0ptCJp0xgdabdcpVtCB5"
            for i, cut in enumerate(cuts):
                if should_stop():
                    set_pattern_state(pat_key, status="stopped", message="停止")
                    return
                num = cut["num"]
                aud_path = os.path.join(out_dir, "audio", f"カット{num.zfill(2)}.mp3")
                if os.path.exists(aud_path):
                    set_pattern_state(pat_key, progress=f"音声 {i+1}/{len(cuts)}", message=f"カット{num} スキップ（既存）")
                    continue

                cut_st = get_cut_state(pat_key, num)
                narration = cut_st.get("narration_override") or cut.get("narration", "")
                if not narration:
                    continue

                set_pattern_state(pat_key, progress=f"音声 {i+1}/{len(cuts)}", message=f"カット{num} [ElevenLabs v2] 音声生成中...")
                try:
                    resp = req.post(
                        f"https://api.elevenlabs.io/v1/text-to-speech/{voice_id}",
                        headers={"xi-api-key": elevenlabs_key, "Content-Type": "application/json"},
                        json={
                            "text": narration, "model_id": "eleven_multilingual_v2",
                            "voice_settings": {"stability": 0.6, "similarity_boost": 0.75, "style": 0.15, "use_speaker_boost": True},
                        },
                        timeout=60,
                    )
                    if resp.status_code == 200 and resp.headers.get("content-type", "").startswith("audio"):
                        with open(aud_path, 'wb') as f:
                            f.write(resp.content)
                        set_cut_state(pat_key, num, audio_model="ElevenLabs v2")
                        set_pattern_state(pat_key, message=f"カット{num} ✓ [ElevenLabs v2] ({len(resp.content)//1024}KB)")
                    else:
                        err_detail = resp.text[:150] if resp.status_code != 200 else "音声データなし"
                        set_pattern_state(pat_key, status="error", message=f"カット{num} [ElevenLabs] エラー ({resp.status_code}): {err_detail}")
                        if resp.status_code in (401, 403):
                            return
                except Exception as e:
                    set_pattern_state(pat_key, message=f"カット{num} 音声エラー: {str(e)[:100]}")

        # ===== Step 3: Creatomate 合成 =====
        if "compose" in steps:
            missing_vid, missing_aud = [], []
            for cut in cuts:
                num = cut["num"]
                vp = os.path.join(out_dir, "videos", f"カット{num.zfill(2)}.mp4")
                ap = os.path.join(out_dir, "audio", f"カット{num.zfill(2)}.mp3")
                if not os.path.exists(vp): missing_vid.append(num)
                if not os.path.exists(ap): missing_aud.append(num)
            if missing_vid or missing_aud:
                msg_parts = []
                if missing_vid: msg_parts.append(f"動画なし: カット{','.join(missing_vid)}")
                if missing_aud: msg_parts.append(f"音声なし: カット{','.join(missing_aud)}")
                set_pattern_state(pat_key, status="error",
                                  message=f"合成スキップ — 素材不足: {'; '.join(msg_parts)}")
                return

            set_pattern_state(pat_key, message="Step 3/3: 合成（Creatomate）")

            vid_urls, audio_urls = {}, {}
            for i, cut in enumerate(cuts):
                if should_stop():
                    set_pattern_state(pat_key, status="stopped", message="停止")
                    return
                num = cut["num"]
                vid_path = os.path.join(out_dir, "videos", f"カット{num.zfill(2)}.mp4")
                aud_path = os.path.join(out_dir, "audio", f"カット{num.zfill(2)}.mp3")
                if os.path.exists(vid_path): vid_urls[num] = upload(vid_path, "video/mp4")
                if os.path.exists(aud_path): audio_urls[num] = upload(aud_path, "audio/mpeg")
                set_pattern_state(pat_key, progress=f"アップロード {i+1}/{len(cuts)}", message=f"カット{num}")

            logo_url = ""
            logo_rel = LOGO_MAP.get(pat["school"], "")
            if logo_rel:
                logo_full = os.path.join(ROOT, logo_rel)
                if os.path.exists(logo_full):
                    logo_url = upload(logo_full, "image/jpeg")
                    log.info("Logo uploaded: %s → %s", pat["school"], logo_url)
                else:
                    log.warning("Logo file not found: %s", logo_full)

            se_base = os.path.join(ROOT, "clients/vantan/se/真面目バージョン")
            se_files = {}
            for cat in ['01_impact', '02_negative', '03_neutral', '04_tiktok']:
                cat_dir = os.path.join(se_base, cat)
                if os.path.isdir(cat_dir):
                    se_files[cat] = sorted([f for f in os.listdir(cat_dir) if f.endswith('.mp3')])

            CUT_SE_MAP = {
                '1': '04_tiktok', '2': '02_negative', '3': '02_negative',
                '4': '01_impact', '5': '01_impact',
                '6': '03_neutral', '7': '03_neutral', '8': '03_neutral',
                '9': '01_impact', '10': '01_impact', '11': '04_tiktok',
            }
            se_urls, prev_file = [], None
            for cut in cuts:
                cat = CUT_SE_MAP.get(cut['num'], '03_neutral')
                candidates = se_files.get(cat, [])
                available = [f for f in candidates if f != prev_file] or candidates
                chosen = random.choice(available) if available else None
                se_urls.append(upload(os.path.join(se_base, cat, chosen), "audio/mpeg") if chosen else "")
                prev_file = chosen

            bgm_path = os.path.join(ROOT, "clients/vantan/bgm/01_hopeful/bgm01_piano_Cmaj_70bpm.mp3")
            bgm_url = upload(bgm_path, "audio/mpeg") if os.path.exists(bgm_path) else ""

            set_pattern_state(pat_key, message="Creatomate レンダリング中...")

            elements = []
            for i, cut in enumerate(cuts):
                num = cut['num']
                if num not in vid_urls or num not in audio_urls:
                    continue
                cut_st = get_cut_state(pat_key, num)
                telop_text = cut_st.get("telop_override") or cut['telop']
                scene_elements = [
                    {"type": "video", "source": vid_urls[num], "fit": "cover", "duration": "100%", "loop": True},
                    {"type": "audio", "source": audio_urls[num], "volume": "100%"},
                ]
                if cut['logo'] == '○':
                    if logo_url:
                        scene_elements.append({
                            "type": "image", "source": logo_url,
                            "x": "50%", "y": "50%", "width": "75%", "height": "25%",
                            "fit": "contain", "x_alignment": "50%", "y_alignment": "50%", "z_index": 15,
                        })
                    else:
                        # ロゴ画像がない場合、スクール名をテキストで表示
                        scene_elements.append({
                            "type": "text", "text": pat["school"],
                            "width": "85%", "height": "25%", "x": "50%", "y": "50%",
                            "duration": "100%", "z_index": 15,
                            "fill_color": "#FFFFFF", "font_family": "Noto Sans JP", "font_weight": "900",
                            "shadow_color": "rgba(0,0,0,0.7)", "shadow_blur": "30px",
                            "x_alignment": "50%", "y_alignment": "50%", "content_alignment": "center",
                            "dynamic_font_size": True, "font_size_maximum": "80px", "font_size_minimum": "36px",
                            "fit": "shrink",
                        })
                elif telop_text:
                    scene_elements.append({
                        "type": "text", "text": telop_text,
                        "width": "85%", "height": "20%", "x": "50%", "y": "50%",
                        "duration": "100%", "z_index": 15,
                        "fill_color": "#FFFFFF", "font_family": "Noto Sans JP", "font_weight": "900",
                        "shadow_color": "rgba(0,0,0,0.6)", "shadow_blur": "25px",
                        "x_alignment": "50%", "y_alignment": "50%", "content_alignment": "center",
                        "dynamic_font_size": True, "font_size_maximum": "70px", "font_size_minimum": "30px",
                        "fit": "shrink",
                    })
                if i < len(se_urls) and se_urls[i]:
                    scene_elements.append({"type": "audio", "source": se_urls[i], "volume": "30%", "duration": "100%"})
                scene = {"type": "composition", "track": 1, "elements": scene_elements}
                if len(elements) > 0:
                    scene["transition"] = {"type": "crossfade", "duration": 0.1}
                elements.append(scene)

            if bgm_url:
                elements.append({"type": "audio", "source": bgm_url, "track": 2, "volume": "30%", "duration": "100%"})

            if not elements:
                set_pattern_state(pat_key, status="error", message="合成する素材がありません")
                return

            cr_resp = req.post(
                "https://api.creatomate.com/v1/renders",
                headers={"Authorization": f"Bearer {creatomate_key}", "Content-Type": "application/json"},
                json={"source": {"output_format": "mp4", "frame_rate": 30, "width": 720, "height": 1280, "elements": elements}},
                timeout=60,
            )
            renders = cr_resp.json()
            render_obj = renders[0] if isinstance(renders, list) else renders
            render_id = render_obj.get("id", "")

            for poll_i in range(60):
                if should_stop():
                    set_pattern_state(pat_key, status="stopped", message="停止")
                    return
                time.sleep(10)
                poll = req.get(f"https://api.creatomate.com/v1/renders/{render_id}",
                              headers={"Authorization": f"Bearer {creatomate_key}"}, timeout=30)
                status = poll.json().get("status", "")
                set_pattern_state(pat_key, message=f"レンダリング: {status} ({poll_i+1}/60)")
                if status == "succeeded":
                    final_url = poll.json().get("url", "")
                    vid = req.get(final_url, timeout=120)
                    final_path = os.path.join(out_dir, "final.mp4")
                    with open(final_path, 'wb') as f:
                        f.write(vid.content)
                    set_pattern_state(pat_key, status="done", progress=f"{len(cuts)}/{len(cuts)}",
                                     message=f"完成 ({len(vid.content)//1024}KB)")
                    return
                elif status == "failed":
                    set_pattern_state(pat_key, status="error", message=f"失敗: {poll.json().get('error_message','')}")
                    return

            set_pattern_state(pat_key, status="error", message="タイムアウト（10分）")
            return

        if "compose" not in steps:
            set_pattern_state(pat_key, status="done", message="素材生成完了（合成はまだ）")

    except Exception as e:
        set_pattern_state(pat_key, status="error", message=str(e)[:200])

# ============================================================
# FastAPI
# ============================================================
app = FastAPI()

# --- カートリッジ API（XLSX ベース） ---

@app.get("/api/cartridges")
def api_cartridges():
    """briefs/ 内の XLSX 一覧 + レジストリ情報 + 現在ロード中"""
    reg = load_registry()
    meta = load_meta()
    files = []
    for f in sorted(os.listdir(BRIEFS_DIR)):
        if not f.endswith(".xlsx"):
            continue
        fpath = os.path.join(BRIEFS_DIR, f)
        stat = os.stat(fpath)
        # レジストリから CID を逆引き
        cid = ""
        for c, info in reg.get("cartridges", {}).items():
            if info["filename"] == f:
                cid = c
                break
        files.append({
            "filename": f,
            "size_kb": round(stat.st_size / 1024, 1),
            "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
            "cid": cid,
        })
    return {
        "files": files,
        "meta": meta,
        "active_cid": meta.get("cartridge_id", ""),
    }

@app.post("/api/cartridges/upload")
async def api_upload(file: UploadFile = File(...)):
    """XLSX ファイルを briefs/ にアップロード"""
    if not file.filename.endswith(".xlsx"):
        return {"error": ".xlsx ファイルのみ対応"}
    dest = os.path.join(BRIEFS_DIR, file.filename)
    content = await file.read()
    with open(dest, "wb") as f:
        f.write(content)
    return {"ok": True, "filename": file.filename, "size_kb": round(len(content) / 1024, 1)}

@app.post("/api/cartridges/load")
def api_load_cartridge(data: dict):
    """XLSX をパースしてカートリッジとして登録・アクティブ化"""
    filename = data.get("filename", "")
    fpath = os.path.join(BRIEFS_DIR, filename)
    if not os.path.exists(fpath):
        return {"error": f"ファイルが見つかりません: {filename}"}

    try:
        patterns = parse_xlsx_to_patterns(fpath)
        if not patterns:
            return {"error": "パターンが見つかりません"}

        cid = register_cartridge(filename)
        save_cartridge_patterns(cid, patterns)

        # 出力ディレクトリ: workflow_002 のシンボリックリンクを自動作成
        output_dest = cartridge_output_base(cid)
        legacy = os.path.join(VANTAN, "output", "workflow_002")
        if os.path.isdir(legacy):
            for pat_key in os.listdir(legacy):
                src = os.path.join(legacy, pat_key)
                dst = os.path.join(output_dest, pat_key)
                if os.path.isdir(src) and not os.path.exists(dst):
                    os.symlink(src, dst)

        # 旧 state があれば取り込む（初回のみ）
        sf = _state_path(cid)
        if not os.path.exists(sf):
            old_state_file = os.path.join(ROOT, "control_panel_state.json")
            if os.path.exists(old_state_file):
                shutil.copy2(old_state_file, sf)

        total_cuts = sum(len(p.get("cuts", [])) for p in patterns.values())
        set_active_cid(cid, filename, len(patterns))
        reload_active_patterns()

        schools = sorted(set(p["school"] for p in patterns.values() if p.get("school")))
        return {
            "ok": True, "cid": cid, "filename": filename,
            "patterns": len(patterns), "total_cuts": total_cuts,
            "schools": schools,
        }
    except Exception as e:
        return {"error": str(e)}

@app.post("/api/cartridges/activate/{cid}")
def api_activate_cartridge(cid: str):
    """カートリッジを切り替え"""
    reg = load_registry()
    if cid not in reg.get("cartridges", {}):
        return {"error": f"カートリッジ {cid} が見つかりません"}
    filename = reg["cartridges"][cid]["filename"]
    patterns = load_cartridge_patterns(cid)
    set_active_cid(cid, filename, len(patterns))
    reload_active_patterns()
    return {"ok": True, "cid": cid, "patterns": len(PATTERNS)}

@app.post("/api/cartridges/refresh")
def api_refresh_cartridge():
    """アクティブカートリッジの XLSX を再パース"""
    meta = load_meta()
    cid = meta.get("cartridge_id", "")
    filename = meta.get("loaded_file", "")
    if not cid or not filename:
        return {"error": "アクティブカートリッジなし"}
    fpath = os.path.join(BRIEFS_DIR, filename)
    if not os.path.exists(fpath):
        return {"error": f"ファイルが見つかりません: {filename}"}
    try:
        patterns = parse_xlsx_to_patterns(fpath)
        save_cartridge_patterns(cid, patterns)
        if get_active_cid() == cid:
            reload_active_patterns()
        return {"ok": True, "cid": cid, "patterns": len(patterns)}
    except Exception as e:
        return {"error": str(e)}

# --- 既存 API（カートリッジ対応版） ---

@app.get("/api/status")
def api_status():
    state = load_state()
    cid = get_active_cid()
    result = {}
    for pat_key, pat in PATTERNS.items():
        outputs = check_outputs(pat_key)
        total = len(pat["cuts"])
        pat_state = state.get(pat_key, {"status": "idle", "progress": "", "message": ""})
        cuts_qc = pat_state.get("cuts_qc", {})
        cuts_data = []
        for c in pat["cuts"]:
            cnum = c["num"]
            cqc = cuts_qc.get(str(cnum), {})
            vid_name = f"カット{cnum.zfill(2)}.mp4"
            aud_name = f"カット{cnum.zfill(2)}.mp3"
            cuts_data.append({
                **c,
                "has_video": vid_name in outputs["videos"],
                "has_audio": aud_name in outputs["audio"],
                "qc_status": cqc.get("qc_status", "pending"),
                "narration_override": cqc.get("narration_override"),
                "telop_override": cqc.get("telop_override"),
                "prompt_jp_override": cqc.get("prompt_jp_override"),
                "prompt_en_override": cqc.get("prompt_en_override"),
                "video_versions": count_cut_versions(pat_key, cnum),
                "video_model": cqc.get("video_model", ""),
                "audio_model": cqc.get("audio_model", ""),
            })
        qc_counts = {"pending": 0, "approved": 0, "rejected": 0}
        for cd in cuts_data:
            qc_counts[cd["qc_status"]] = qc_counts.get(cd["qc_status"], 0) + 1
        result[pat_key] = {
            "school": pat["school"], "field": pat["field"], "child": pat["child"],
            "total_cuts": total, "videos": len(outputs["videos"]),
            "audio": len(outputs["audio"]), "has_final": outputs["final"] is not None,
            "cuts": cuts_data, "qc_counts": qc_counts,
            **{k: v for k, v in pat_state.items() if k != "cuts_qc"},
        }
    return {"patterns": result, "active_cid": cid}

@app.post("/api/generate/{pat_key}")
def api_generate(pat_key: str, data: dict = None):
    ps = get_pattern_state(pat_key)
    if ps.get("status") == "running":
        return {"error": "既に実行中です"}
    steps = data.get("steps") if data else None
    video_model = (data or {}).get("video_model", "fal-ai/veo3.1")
    if pat_key in PATTERNS:
        PATTERNS[pat_key]["video_model"] = video_model
    set_pattern_state(pat_key, status="running", progress="0/?", message=f"開始中... [{video_model}]", stop_requested=False)
    t = threading.Thread(target=run_pipeline, args=(pat_key, steps), daemon=True)
    t.start()
    return {"ok": True, "pattern": pat_key, "steps": steps, "video_model": video_model}

@app.post("/api/stop/{pat_key}")
def api_stop(pat_key: str):
    set_pattern_state(pat_key, stop_requested=True, message="停止リクエスト中...")
    return {"ok": True}

@app.post("/api/regenerate/{pat_key}")
def api_regenerate(pat_key: str, data: dict = None):
    ps = get_pattern_state(pat_key)
    if ps.get("status") == "running":
        return {"error": "既に実行中です"}
    target = (data or {}).get("target", "all")  # "video", "audio", "all"
    out_dir = os.path.join(get_output_base(), pat_key)

    if target in ("video", "all"):
        d = os.path.join(out_dir, "videos")
        if os.path.isdir(d):
            shutil.rmtree(d)
    if target in ("audio", "all"):
        d = os.path.join(out_dir, "audio")
        if os.path.isdir(d):
            shutil.rmtree(d)
    # final は動画か音声どちらかが変わったら消す
    final = os.path.join(out_dir, "final.mp4")
    if os.path.exists(final):
        os.remove(final)

    steps_map = {"video": ["video"], "audio": ["narration"], "all": ["video", "narration"]}
    steps = steps_map.get(target, ["video", "narration"])
    label_map = {"video": "動画のみ再生成", "audio": "音声のみ再生成", "all": "全再生成"}
    label = label_map.get(target, "再生成")

    set_pattern_state(pat_key, status="running", progress="0/?", message=label, stop_requested=False)
    t = threading.Thread(target=run_pipeline, args=(pat_key, steps), daemon=True)
    t.start()
    return {"ok": True, "pattern": pat_key, "target": target}

@app.post("/api/generate-batch")
def api_generate_batch(data: dict):
    pat_keys = data.get("patterns", [])
    steps = data.get("steps")
    video_model = data.get("video_model", "fal-ai/veo3.1")
    def run_batch():
        for pk in pat_keys:
            s = load_state()
            if s.get(pk, {}).get("stop_requested"):
                break
            # video_model をパターンに一時設定
            if pk in PATTERNS:
                PATTERNS[pk]["video_model"] = video_model
            run_pipeline(pk, steps)
    for pk in pat_keys:
        set_pattern_state(pk, status="queued", message=f"バッチ待機中... [{video_model}]", stop_requested=False)
    if pat_keys:
        set_pattern_state(pat_keys[0], status="running", message=f"バッチ開始... [{video_model}]")
    t = threading.Thread(target=run_batch, daemon=True)
    t.start()
    return {"ok": True, "patterns": pat_keys, "video_model": video_model}

@app.post("/api/regenerate-cut/{pat_key}/{cut_num}")
def api_regenerate_cut(pat_key: str, cut_num: str):
    ps = get_pattern_state(pat_key)
    if ps.get("status") == "running":
        return {"error": "既に実行中です"}
    out_dir = os.path.join(get_output_base(), pat_key)
    for ext, sub in [(".mp4", "videos"), (".mp3", "audio")]:
        path = os.path.join(out_dir, sub, f"カット{cut_num.zfill(2)}{ext}")
        if os.path.exists(path):
            ver = count_cut_versions(pat_key, cut_num)
            archive_name = f"カット{cut_num.zfill(2)}_v{str(ver).zfill(3)}{ext}"
            os.rename(path, os.path.join(out_dir, sub, archive_name))
    final = os.path.join(out_dir, "final.mp4")
    if os.path.exists(final):
        os.remove(final)
    set_pattern_state(pat_key, status="running", progress=f"カット{cut_num}再生成",
                     message=f"カット{cut_num}を再生成中...", stop_requested=False)
    t = threading.Thread(target=run_pipeline, args=(pat_key, ["video", "narration"]), daemon=True)
    t.start()
    return {"ok": True, "pattern": pat_key, "cut": cut_num}

@app.post("/api/reload-sheets")
def api_reload_sheets():
    """アクティブカートリッジの XLSX を再パース"""
    return api_refresh_cartridge()

@app.post("/api/stop-all")
def api_stop_all():
    state = load_state()
    for k in state:
        state[k]["stop_requested"] = True
    save_state(state)
    return {"ok": True}

# --- QC・テキスト編集 API ---

@app.put("/api/cut/{pat_key}/{cut_num}/status")
def api_cut_status(pat_key: str, cut_num: str, data: dict):
    new_status = data.get("qc_status", "pending")
    if new_status not in ("pending", "approved", "rejected"):
        return {"error": "Invalid status"}
    set_cut_state(pat_key, cut_num, qc_status=new_status)
    return {"ok": True}

@app.put("/api/cut/{pat_key}/{cut_num}/text")
def api_cut_text(pat_key: str, cut_num: str, data: dict):
    pat = PATTERNS.get(pat_key)
    if not pat:
        return {"error": "Pattern not found"}
    cut = next((c for c in pat["cuts"] if c["num"] == cut_num), None)
    if not cut:
        return {"error": "Cut not found"}
    reason = data.get("reason", "")
    updates = {}
    for field in ["narration", "telop", "prompt_jp", "prompt_en"]:
        if field in data and data[field] is not None:
            override_key = f"{field}_override"
            original = cut.get(field, "")
            new_val = data[field]
            if new_val != original:
                updates[override_key] = new_val
                log_correction(pat_key, cut_num, field, original, new_val, reason)
            else:
                updates[override_key] = None
    if updates:
        set_cut_state(pat_key, cut_num, **updates)
    return {"ok": True}

@app.get("/api/cut/{pat_key}/{cut_num}/versions")
def api_cut_versions(pat_key: str, cut_num: str):
    videos = list_cut_versions(pat_key, cut_num)
    cid = get_active_cid()
    base = os.path.join(get_output_base(), pat_key, "audio")
    audio = []
    if os.path.isdir(base):
        prefix = f"カット{cut_num.zfill(2)}"
        files = sorted([f for f in os.listdir(base) if f.startswith(prefix) and f.endswith('.mp3')])
        audio = [{"name": f, "url": f"/media/{cid}/{pat_key}/audio/{f}"} for f in files]
    return {"videos": videos, "audio": audio}

@app.get("/api/corrections/{pat_key}")
def api_corrections(pat_key: str):
    cid = get_active_cid()
    if not cid:
        return []
    path = _corrections_path(cid)
    if os.path.exists(path):
        with open(path, encoding="utf-8") as f:
            entries = json.load(f)
        return [e for e in entries if e.get("pattern") == pat_key]
    return []

@app.put("/api/pattern/{pat_key}/approve-all")
def api_approve_all(pat_key: str):
    pat = PATTERNS.get(pat_key)
    if not pat:
        return {"error": "Pattern not found"}
    for c in pat["cuts"]:
        set_cut_state(pat_key, c["num"], qc_status="approved")
    return {"ok": True}

@app.put("/api/pattern/{pat_key}/pending-all")
def api_pending_all(pat_key: str):
    pat = PATTERNS.get(pat_key)
    if not pat:
        return {"error": "Pattern not found"}
    for c in pat["cuts"]:
        set_cut_state(pat_key, c["num"], qc_status="pending")
    return {"ok": True}

# --- ファイル配信 ---

@app.get("/api/files/{pat_key}")
def api_files(pat_key: str):
    cid = get_active_cid()
    base = os.path.join(get_output_base(), pat_key)
    result = {"videos": [], "audio": [], "final": None}
    vid_dir, aud_dir = os.path.join(base, "videos"), os.path.join(base, "audio")
    if os.path.isdir(vid_dir):
        for f in sorted(os.listdir(vid_dir)):
            if f.endswith('.mp4'):
                result["videos"].append({"name": f, "url": f"/media/{cid}/{pat_key}/videos/{f}"})
    if os.path.isdir(aud_dir):
        for f in sorted(os.listdir(aud_dir)):
            if f.endswith('.mp3'):
                result["audio"].append({"name": f, "url": f"/media/{cid}/{pat_key}/audio/{f}"})
    final = os.path.join(base, "final.mp4")
    if os.path.exists(final):
        result["final"] = f"/media/{cid}/{pat_key}/final.mp4"
    return result

# メディア配信（output ディレクトリ全体）
output_dir = os.path.join(VANTAN, "output")
if os.path.isdir(output_dir):
    app.mount("/media", StaticFiles(directory=output_dir, follow_symlink=True), name="media")

assets_dir = os.path.join(ROOT, "assets")
if os.path.isdir(assets_dir):
    app.mount("/assets", StaticFiles(directory=assets_dir), name="assets")

@app.get("/", response_class=HTMLResponse)
def index():
    html_path = os.path.join(ROOT, "control_panel_ui.html")
    with open(html_path, encoding="utf-8") as f:
        content = f.read()
    return Response(content=content, media_type="text/html", headers={
        "Cache-Control": "no-cache, no-store, must-revalidate",
        "Pragma": "no-cache", "Expires": "0",
    })

if __name__ == "__main__":
    print(f"\n  VANTAN Video Studio v1.2: http://localhost:8888\n")
    uvicorn.run(app, host="0.0.0.0", port=8888)

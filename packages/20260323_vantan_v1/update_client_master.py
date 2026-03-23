"""
クライアントマスタースプレッドシート更新スクリプト

clients/vantan/ 内のExcelファイルを読み込み、
Google Sheetsのマスタースプレッドシートを自動更新する。

使い方:
    python3 update_client_master.py
"""

import openpyxl
import gspread
import glob
import os
from datetime import datetime

# === 設定 ===
CLIENT_DIR = "clients/vantan"
MASTER_SHEET_NAME = "VANTAN_クライアントマスター"

# === Excel読み込み ===
def load_excel_data(client_dir):
    """クライアントフォルダ内のExcelファイルを読み込む"""
    xlsx_files = glob.glob(os.path.join(client_dir, "*.xlsx"))
    if not xlsx_files:
        print("Excelファイルが見つかりません")
        return None

    print(f"読み込み: {xlsx_files[0]}")
    wb = openpyxl.load_workbook(xlsx_files[0])
    return wb


def extract_school_data(wb):
    """各スクール訴求内容シートからデータを抽出"""
    ws = wb["各スクール訴求内容"]

    schools = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = row[2]   # スクールコード
        name = row[3]   # スクール名称
        dept = row[4]   # 部
        course = row[5] # コース
        target = row[6] # ターゲット（本人向け/親向け）
        demo = row[7]   # デモグラ
        overview = row[8] or ""  # スクール概要
        lp = row[9] or ""   # LP URL
        hp = row[10] or ""  # HP URL

        if not code:
            continue

        if code not in schools:
            schools[code] = {
                "name": name or "",
                "courses": [],
            }

        schools[code]["courses"].append({
            "dept": dept or "",
            "course": course or "",
            "target": target or "",
            "demo": demo or "",
            "overview": overview,
            "lp": lp,
            "hp": hp,
        })

    return schools


def extract_regulations(wb):
    """依頼内容まとめシートからNGレギュレーション等を抽出"""
    ws = wb["202626依頼内容まとめ"]
    lines = []
    for row in ws.iter_rows(values_only=True):
        val = row[0]
        if val:
            lines.append(str(val).strip())
    return "\n".join(lines)


def build_master_data(schools, regulations):
    """マスタースプレッドシート用のデータを構築"""
    data = [
        ["カテゴリ", "項目", "値", "補足"],

        # --- 基本情報 ---
        ["基本情報", "クライアント名", "株式会社バンタン", "KADOKAWAグループ"],
        ["基本情報", "ブランド群", "バンタンデザイン研究所、バンタンゲームアカデミー、ヴィーナスアカデミー 他13スクール", ""],
        ["基本情報", "所在地", "東京都渋谷区／大阪府大阪市", "東京校・大阪校"],
        ["基本情報", "種別", "専門スクール（認可校ではなくスクール）", "学歴は「専門学校卒」にはならない"],
        ["基本情報", "設立", "1965年", "60年以上の歴史"],

        # --- USP ---
        ["USP", "全体USP①", "100%クリエイター教育：全授業が現役プロによる指導", "最大の差別化ポイント"],
        ["USP", "全体USP②", "企業連携カリキュラム：実案件に取り組む実践型", ""],
        ["USP", "全体USP③", "KADOKAWAグループの業界ネットワーク", "就職・デビュー支援に直結"],
        ["USP", "全体USP④", "少人数制クラス", ""],
        ["USP", "全体USP⑤", "渋谷・大阪の都心立地", ""],

        # --- NGレギュレーション ---
        ["レギュレーション", "NG①", "エンドユーザー目線で炎上リスクがあるもの", "例：メイク道具を雑に扱う、漫画の学校なのにAI素材を使う等"],
        ["レギュレーション", "NG②", "遷移先LPと著しく内容や雰囲気が逸脱したもの", "親向けLPに本人向け内容、トンマナの逸脱"],
        ["レギュレーション", "NG③", "新規バナーなのに他LPと一緒の構成・一部素材替え", "UGCとアニメバージョンくらいのバリエーション必要"],
        ["レギュレーション", "NG④", "注釈・スクール名の表記誤り", "レギュレーションシートの確認必須"],
        ["レギュレーション", "NG⑤", "「No.1」「最高」等の最上級表現は根拠なく使用不可", "景品表示法に注意"],

        # --- 配信情報 ---
        ["配信情報", "案件名", "株式会社バンタン_バンタン_UPPA_ダイレクト配信", ""],
        ["配信情報", "配信媒体", "Meta（Facebook/Instagram）", ""],
        ["配信情報", "動画フォーマット", "縦型9:16", ""],

        # --- セパレータ ---
        ["", "", "", ""],
        ["スクール一覧", "---", "--- 以下、スクール別情報 ---", ""],
    ]

    # --- スクール別情報 ---
    for code, school in schools.items():
        school_name = school["name"]
        # スクールヘッダー
        data.append(["", "", "", ""])
        data.append([f"【{code}】{school_name}", "スクール名", school_name, f"コード: {code}"])

        # コースごとの情報をまとめる
        depts = {}
        for c in school["courses"]:
            dept_key = c["dept"]
            if dept_key not in depts:
                depts[dept_key] = []
            depts[dept_key].append(c)

        for dept_name, courses in depts.items():
            # 部の情報
            course_names = [c["course"] for c in courses if c["course"]]
            target_types = list(set(c["target"] for c in courses if c["target"]))
            demos = list(set(c["demo"] for c in courses if c["demo"]))

            cat = f"【{code}】{school_name}"

            data.append([cat, f"部: {dept_name}", "", ""])

            if course_names:
                data.append([cat, "コース", "、".join(course_names), ""])

            if target_types:
                data.append([cat, "ターゲット種別", "、".join(target_types), f"部: {dept_name}"])

            if demos:
                data.append([cat, "デモグラ", "、".join(demos), f"部: {dept_name}"])

            # LP/HP（最初に見つかったもの）
            for c in courses:
                if c["lp"]:
                    # LPが長い概要テキストの場合はスキップ
                    lp_text = str(c["lp"])
                    if lp_text.startswith("http"):
                        data.append([cat, "LP", lp_text, f"部: {dept_name}"])
                    break

            for c in courses:
                if c["hp"] and str(c["hp"]).startswith("http"):
                    data.append([cat, "HP", str(c["hp"]), f"部: {dept_name}"])
                    break

            # スクール概要（長いテキストがあるもの）
            for c in courses:
                overview = str(c["overview"]).strip()
                if len(overview) > 50:  # 短いデモグラ情報ではなく実際の概要
                    # 概要が非常に長い場合は先頭500文字に切り詰め
                    if len(overview) > 500:
                        overview = overview[:500] + "..."
                    course_label = c["course"] if c["course"] else dept_name
                    data.append([cat, f"概要: {course_label}", overview, ""])

    # --- メタ情報 ---
    data.append(["", "", "", ""])
    data.append(["メタ情報", "最終更新日", datetime.now().strftime("%Y-%m-%d %H:%M"), "自動更新"])
    data.append(["メタ情報", "データソース", "【バンタン_CA極案件】各スクール訴求内容.xlsx", ""])
    data.append(["メタ情報", "次回更新予定", "", "キャンペーン切替時・新コース追加時"])

    return data


# === スプレッドシート更新 ===
def update_spreadsheet(data):
    """マスタースプレッドシートを更新"""
    gc = gspread.oauth(
        credentials_filename="oauth_credentials.json",
        authorized_user_filename="token.json",
    )

    # 既存のマスタースプレッドシートを探す
    try:
        sh = gc.open(MASTER_SHEET_NAME)
        print(f"既存のスプレッドシートを更新: {MASTER_SHEET_NAME}")
    except gspread.SpreadsheetNotFound:
        sh = gc.create(MASTER_SHEET_NAME)
        print(f"新規作成: {MASTER_SHEET_NAME}")

    ws = sh.sheet1
    ws.update_title("クライアント情報")

    # 全データクリアして書き直し
    ws.clear()
    ws.update(values=data, range_name="A1")

    # 書式設定
    ws.format("A1:D1", {
        "textFormat": {"bold": True, "foregroundColorStyle": {"rgbColor": {"red": 1, "green": 1, "blue": 1}}},
        "backgroundColor": {"red": 0.2, "green": 0.3, "blue": 0.5},
    })
    ws.format("A:A", {"textFormat": {"bold": True}})
    ws.format("A:D", {"wrapStrategy": "WRAP"})

    # 列幅
    reqs = [
        {"updateDimensionProperties": {"range": {"sheetId": ws.id, "dimension": "COLUMNS", "startIndex": 0, "endIndex": 1}, "properties": {"pixelSize": 200}, "fields": "pixelSize"}},
        {"updateDimensionProperties": {"range": {"sheetId": ws.id, "dimension": "COLUMNS", "startIndex": 1, "endIndex": 2}, "properties": {"pixelSize": 180}, "fields": "pixelSize"}},
        {"updateDimensionProperties": {"range": {"sheetId": ws.id, "dimension": "COLUMNS", "startIndex": 2, "endIndex": 3}, "properties": {"pixelSize": 500}, "fields": "pixelSize"}},
        {"updateDimensionProperties": {"range": {"sheetId": ws.id, "dimension": "COLUMNS", "startIndex": 3, "endIndex": 4}, "properties": {"pixelSize": 250}, "fields": "pixelSize"}},
    ]
    sh.batch_update({"requests": reqs})

    print(f"更新完了! {len(data)}行")
    print(f"URL: {sh.url}")
    return sh.url


# === メイン ===
if __name__ == "__main__":
    wb = load_excel_data(CLIENT_DIR)
    if wb is None:
        exit(1)

    schools = extract_school_data(wb)
    regulations = extract_regulations(wb)

    print(f"スクール数: {len(schools)}")
    print(f"レギュレーション: {len(regulations)}文字")

    data = build_master_data(schools, regulations)
    url = update_spreadsheet(data)

"""QC Gallery App v2 - Recruit Agent ad creative QC with cartridge-style brief loading.

Cartridge = xlsx設計図ファイル。briefs/ フォルダに置くと自動でプルダウンに表示。
選択 → インポート → 画像生成 → QC → デザイナー納品 の一気通貫フロー。
"""

from __future__ import annotations

import json
import os
import shutil
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Optional

import fal_client
import httpx
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.responses import HTMLResponse
from fastapi.staticfiles import StaticFiles
from google import genai
from openpyxl import load_workbook
from pydantic import BaseModel

load_dotenv(Path(__file__).parent / ".env")

APP_DIR = Path(__file__).parent
IMAGES_DIR = APP_DIR / "images"
BRIEFS_DIR = APP_DIR / "briefs"
STATE_DIR = APP_DIR / "state"
META_PATH = APP_DIR / "cartridge_meta.json"
PROMPT_CACHE_PATH = APP_DIR / "prompt_cache.json"
FAL_KEY = os.getenv("FAL_KEY", "")
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "")

IMAGES_DIR.mkdir(exist_ok=True)
BRIEFS_DIR.mkdir(exist_ok=True)
STATE_DIR.mkdir(exist_ok=True)

REGISTRY_PATH = STATE_DIR / "cartridge_registry.json"


# ---------------------------------------------------------------------------
# Cartridge registry: C001, C002, ... → xlsx filename mapping
# ---------------------------------------------------------------------------

def load_registry() -> dict:
    """Load cartridge registry: {"next_id": N, "cartridges": {"C001": {...}, ...}}"""
    if REGISTRY_PATH.exists():
        with open(REGISTRY_PATH, encoding="utf-8") as f:
            return json.load(f)
    return {"next_id": 1, "cartridges": {}}


def save_registry(reg: dict) -> None:
    with open(REGISTRY_PATH, "w", encoding="utf-8") as f:
        json.dump(reg, f, ensure_ascii=False, indent=2)


def register_cartridge(filename: str) -> str:
    """Register an xlsx file and return its cartridge ID (e.g. C001).

    If the same filename was already registered, return the existing ID.
    """
    reg = load_registry()

    # Check if already registered
    for cid, info in reg["cartridges"].items():
        if info["filename"] == filename:
            return cid

    # Assign new ID
    num = reg["next_id"]
    cid = f"C{num:03d}"
    reg["cartridges"][cid] = {
        "filename": filename,
        "registered_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    reg["next_id"] = num + 1
    save_registry(reg)
    return cid


def get_cartridge_filename(cid: str) -> str | None:
    """Look up the original xlsx filename for a cartridge ID."""
    reg = load_registry()
    info = reg["cartridges"].get(cid)
    return info["filename"] if info else None


def _cartridge_id() -> str:
    """Return current cartridge ID from meta, or empty string."""
    meta = load_meta()
    return meta.get("cartridge_id", "")


def _briefs_path(cid: str = "") -> Path:
    cid = cid or _cartridge_id()
    if not cid:
        return APP_DIR / "briefs.json"
    return STATE_DIR / f"{cid}_briefs.json"


def _state_path(cid: str = "") -> Path:
    cid = cid or _cartridge_id()
    if not cid:
        return APP_DIR / "gallery_state.json"
    return STATE_DIR / f"{cid}_state.json"


def _images_dir(cid: str = "") -> Path:
    cid = cid or _cartridge_id()
    if not cid:
        return IMAGES_DIR
    d = IMAGES_DIR / cid
    d.mkdir(exist_ok=True)
    return d

app = FastAPI(title="Recruit Agent QC Gallery v2")

# Shared kage-lab theme (repo root assets/)
REPO_ROOT = APP_DIR.parent.parent
_ASSETS_DIR = REPO_ROOT / "assets"
if _ASSETS_DIR.is_dir():
    app.mount("/assets", StaticFiles(directory=str(_ASSETS_DIR)), name="kl_assets")

# Serve versioned images
app.mount("/images", StaticFiles(directory=str(IMAGES_DIR)), name="images")


# ---------------------------------------------------------------------------
# Column mapping: xlsx header → briefs.json key
# ---------------------------------------------------------------------------
COLUMN_MAP = {
    "No.": "num",
    "業界セグメント": "segment",
    "訴求軸": "concept",
    "ターゲット心理（インサイト）": "insight",
    "ターゲット心理": "insight",
    "コピー HL（メインコピー）": "copy_hl",
    "コピーHL（メインコピー）": "copy_hl",
    "コピーHL": "copy_hl",
    "コピー Body（サブテキスト）": "copy_body",
    "コピーBody（サブテキスト）": "copy_body",
    "コピーBody": "copy_body",
    "CTA": "copy_cta",
    "ビジュアル方向性": "visual_direction",
    "人物設定": "character",
    "色・トーン": "tone",
    "画像プロンプト（日本語）": "prompt",
    "画像プロンプト": "prompt",
    "注釈": "note",
    # English fallbacks
    "segment": "segment",
    "concept": "concept",
    "insight": "insight",
    "copy_hl": "copy_hl",
    "copy_body": "copy_body",
    "copy_cta": "copy_cta",
    "visual_direction": "visual_direction",
    "character": "character",
    "tone": "tone",
    "prompt": "prompt",
}


# ---------------------------------------------------------------------------
# Data helpers
# ---------------------------------------------------------------------------

def load_briefs(cid: str = "") -> list[dict]:
    p = _briefs_path(cid)
    if not p.exists():
        return []
    with open(p, encoding="utf-8") as f:
        return json.load(f)


def save_briefs(briefs: list[dict], cid: str = "") -> None:
    p = _briefs_path(cid)
    with open(p, "w", encoding="utf-8") as f:
        json.dump(briefs, f, ensure_ascii=False, indent=2)


def load_state(cid: str = "") -> dict:
    p = _state_path(cid)
    if p.exists():
        with open(p, encoding="utf-8") as f:
            return json.load(f)
    briefs = load_briefs(cid)
    state = {}
    for b in briefs:
        state[str(b["num"])] = {
            "qc_status": "pending",
            "selected_version": None,
            "version_count": 0,
        }
    save_state(state, cid)
    return state


def save_state(state: dict, cid: str = "") -> None:
    p = _state_path(cid)
    fd, tmp = tempfile.mkstemp(dir=STATE_DIR, suffix=".tmp")
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            json.dump(state, f, ensure_ascii=False, indent=2)
        os.replace(tmp, p)
    except Exception:
        os.unlink(tmp)
        raise


def load_meta() -> dict:
    if META_PATH.exists():
        with open(META_PATH, encoding="utf-8") as f:
            return json.load(f)
    return {"loaded_file": None, "loaded_at": None, "brief_count": 0}


def save_meta(meta: dict) -> None:
    with open(META_PATH, "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)


def get_current_image(num: int, entry: dict, cid: str = "") -> str | None:
    vc = entry["version_count"]
    if vc == 0:
        return None
    sel = entry["selected_version"]
    v = sel if sel is not None else vc
    cid = cid or _cartridge_id()
    if cid:
        return f"/images/{cid}/{num:02d}/v{v:03d}.jpg"
    return f"/images/{num:02d}/v{v:03d}.jpg"


def check_fal_key() -> None:
    if not FAL_KEY:
        raise HTTPException(status_code=500, detail="FAL_KEY not set in .env")


# ---------------------------------------------------------------------------
# XLSX → briefs parser
# ---------------------------------------------------------------------------

def parse_xlsx_to_briefs(filepath: Path) -> list[dict]:
    """Parse an xlsx file into a list of brief dicts."""
    wb = load_workbook(filepath, read_only=True, data_only=True)
    # Use first sheet (or the one named "制作設計ブリーフ*")
    ws = None
    for name in wb.sheetnames:
        if "ブリーフ" in name or "brief" in name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        raise ValueError("スプレッドシートにデータ行がありません")

    # Map headers
    raw_headers = [str(h).strip() if h else "" for h in rows[0]]
    col_keys = []
    for h in raw_headers:
        matched = COLUMN_MAP.get(h)
        if not matched:
            # Fuzzy match: check if header contains a known key
            for pattern, key in COLUMN_MAP.items():
                if pattern in h:
                    matched = key
                    break
        col_keys.append(matched)

    briefs = []
    for row_idx, row in enumerate(rows[1:], start=1):
        brief = {}
        for col_idx, val in enumerate(row):
            if col_idx >= len(col_keys) or col_keys[col_idx] is None:
                continue
            key = col_keys[col_idx]
            if val is None:
                val = ""
            if key == "num":
                try:
                    val = int(val)
                except (ValueError, TypeError):
                    val = row_idx
            else:
                val = str(val).strip()
            brief[key] = val

        # Skip empty rows
        if not brief.get("segment") and not brief.get("copy_hl"):
            continue

        # Ensure num exists
        if "num" not in brief:
            brief["num"] = row_idx

        # Fill defaults
        for field in ["segment", "concept", "insight", "copy_hl", "copy_body",
                       "copy_cta", "visual_direction", "character", "tone", "prompt"]:
            brief.setdefault(field, "")

        briefs.append(brief)

    return briefs


# ---------------------------------------------------------------------------
# Prompt translation (JA → EN) with Flux Pro optimization
# ---------------------------------------------------------------------------

TRANSLATE_SYSTEM_PROMPT = """\
You are a prompt translator for Flux Pro (fal.ai) image generation.

Your job: translate the Japanese image-generation prompt into an English prompt
optimized for Flux Pro.

Rules:
1. Translate faithfully — do NOT add concepts that are not in the original.
2. Remove any salary/income numbers (e.g. 年収780万) — these cause text artifacts.
3. Remove any text that says "年収交渉" or similar text-on-object instructions.
4. Keep "no text, no logos, no signage" instruction.
5. Structure: Subject → Pose/Action → Environment → Lighting → Style/Technical.
6. Use natural English sentences, not keyword lists.
7. Add camera/lens hint for realism: e.g. "shot on 85mm lens, f/2.8"
8. Keep it 40-80 words. Do not exceed 100 words.
9. Output ONLY the English prompt, nothing else.
"""


def load_prompt_cache() -> dict:
    if PROMPT_CACHE_PATH.exists():
        with open(PROMPT_CACHE_PATH, encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_prompt_cache(cache: dict) -> None:
    with open(PROMPT_CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


def translate_prompt(ja_prompt: str) -> str:
    cache = load_prompt_cache()
    if ja_prompt in cache:
        return cache[ja_prompt]

    if not GEMINI_API_KEY:
        raise HTTPException(
            status_code=500,
            detail="GEMINI_API_KEY not set — needed for prompt translation",
        )

    client = genai.Client(api_key=GEMINI_API_KEY)
    response = client.models.generate_content(
        model="gemini-2.5-flash",
        contents=ja_prompt,
        config={"system_instruction": TRANSLATE_SYSTEM_PROMPT},
    )
    en_prompt = response.text.strip()

    cache[ja_prompt] = en_prompt
    save_prompt_cache(cache)
    return en_prompt


# ---------------------------------------------------------------------------
# Endpoints — Cartridge Management
# ---------------------------------------------------------------------------

@app.get("/", response_class=HTMLResponse)
async def index():
    with open(APP_DIR / "gallery.html", encoding="utf-8") as f:
        return f.read()


@app.get("/api/cartridges")
async def list_cartridges():
    """List available xlsx files in briefs/ directory."""
    reg = load_registry()
    # Reverse lookup: filename → cid
    file_to_cid = {info["filename"]: cid for cid, info in reg["cartridges"].items()}

    files = []
    for f in sorted(BRIEFS_DIR.glob("*.xlsx")):
        stat = f.stat()
        cid = file_to_cid.get(f.name)
        files.append({
            "filename": f.name,
            "cartridge_id": cid,
            "size_kb": round(stat.st_size / 1024, 1),
            "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
        })
    # Also check root for legacy xlsx files
    for f in sorted(APP_DIR.glob("*.xlsx")):
        stat = f.stat()
        display = f"(root) {f.name}"
        cid = file_to_cid.get(f.name)
        files.append({
            "filename": display,
            "path": f.name,
            "cartridge_id": cid,
            "size_kb": round(stat.st_size / 1024, 1),
            "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
        })
    meta = load_meta()
    return {"files": files, "meta": meta, "registry": reg["cartridges"]}


@app.post("/api/cartridges/upload")
async def upload_cartridge(file: UploadFile = File(...)):
    """Upload an xlsx file to the briefs/ directory."""
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="xlsxファイルのみアップロード可能です")

    dest = BRIEFS_DIR / file.filename
    content = await file.read()
    dest.write_bytes(content)

    return {"ok": True, "filename": file.filename, "size_kb": round(len(content) / 1024, 1)}


@app.post("/api/cartridges/load")
async def load_cartridge(body: dict):
    """Load (import) an xlsx cartridge → parse → save briefs + init/restore state."""
    filename = body.get("filename", "")
    if not filename:
        raise HTTPException(status_code=400, detail="filename is required")

    # Resolve path: check briefs/ first, then root
    filepath = BRIEFS_DIR / filename
    if not filepath.exists():
        filepath = APP_DIR / filename
    # Handle "(root) " prefix from legacy files
    if not filepath.exists() and filename.startswith("(root) "):
        filepath = APP_DIR / filename[7:]
    if not filepath.exists():
        raise HTTPException(status_code=404, detail=f"File not found: {filename}")

    try:
        briefs = parse_xlsx_to_briefs(filepath)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"XLSX parse error: {e}")

    if not briefs:
        raise HTTPException(status_code=400, detail="ブリーフが見つかりませんでした")

    # Register cartridge and get ID (C001, C002, ...)
    display_name = filename.replace("(root) ", "")
    cid = register_cartridge(display_name)

    # Save briefs for this cartridge
    save_briefs(briefs, cid)

    # Build state: scan only this cartridge's image folder for existing versions
    img_base = _images_dir(cid)
    state = {}
    for b in briefs:
        key = str(b["num"])
        num_dir = img_base / f"{b['num']:02d}"
        vc = 0
        if num_dir.exists():
            vc = len(list(num_dir.glob("v*.jpg")))
        state[key] = {
            "qc_status": "pending",
            "selected_version": None,
            "version_count": vc,
        }

    # If a state file already exists for this cartridge, restore QC statuses
    existing_state_path = _state_path(cid)
    if existing_state_path.exists():
        with open(existing_state_path, encoding="utf-8") as f:
            old_state = json.load(f)
        for key in state:
            if key in old_state:
                state[key]["qc_status"] = old_state[key].get("qc_status", "pending")
                state[key]["selected_version"] = old_state[key].get("selected_version")
                state[key]["prompt_override"] = old_state[key].get("prompt_override")

    save_state(state, cid)

    # Save meta — cartridge_id is the key link
    meta = {
        "cartridge_id": cid,
        "loaded_file": display_name,
        "loaded_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "brief_count": len(briefs),
    }
    save_meta(meta)

    return {
        "ok": True,
        "filename": display_name,
        "cartridge_id": cid,
        "brief_count": len(briefs),
        "segments": list(set(b["segment"] for b in briefs if b.get("segment"))),
    }


# ---------------------------------------------------------------------------
# Endpoints — Briefs & QC (unchanged logic)
# ---------------------------------------------------------------------------

@app.get("/api/briefs")
async def get_briefs():
    cid = _cartridge_id()
    briefs = load_briefs(cid)
    state = load_state(cid)
    result = []
    for b in briefs:
        key = str(b["num"])
        entry = state.get(key, {"qc_status": "pending", "selected_version": None, "version_count": 0})
        # Apply copy overrides for display
        copy_overrides = entry.get("copy_overrides", {})
        display = {**b}
        for field in ["copy_hl", "copy_body", "copy_cta"]:
            if field in copy_overrides:
                display[field] = copy_overrides[field]

        result.append({
            **display,
            "qc_status": entry["qc_status"],
            "selected_version": entry["selected_version"],
            "version_count": entry["version_count"],
            "current_image": get_current_image(b["num"], entry, cid),
            "prompt_override": entry.get("prompt_override"),
            "copy_overrides": copy_overrides or None,
            "original_copy_hl": b.get("copy_hl", ""),
            "original_copy_body": b.get("copy_body", ""),
            "original_copy_cta": b.get("copy_cta", ""),
            "cartridge_id": cid,
        })
    return result


@app.post("/api/regen/{num}")
async def regen(num: int):
    cid = _cartridge_id()
    briefs = load_briefs(cid)
    brief = next((b for b in briefs if b["num"] == num), None)
    if brief is None:
        raise HTTPException(status_code=404, detail=f"Brief #{num} not found")

    check_fal_key()

    state = load_state(cid)
    key = str(num)
    override = state.get(key, {}).get("prompt_override")
    source_prompt = override if override else brief["prompt"]

    try:
        en_prompt = translate_prompt(source_prompt)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Prompt translation error: {e}")

    try:
        result = fal_client.subscribe(
            "fal-ai/flux-pro/v1.1",
            arguments={
                "prompt": en_prompt,
                "image_size": {"width": 768, "height": 1344},
                "num_images": 1,
                "safety_tolerance": 6,
            },
        )
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"fal.ai API error: {e}")

    if not result.get("images"):
        raise HTTPException(status_code=502, detail="No image returned from fal.ai")

    image_url = result["images"][0]["url"]

    try:
        resp = httpx.get(image_url, timeout=30)
        resp.raise_for_status()
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Image download failed: {e}")

    state = load_state(cid)
    if key not in state:
        state[key] = {"qc_status": "pending", "selected_version": None, "version_count": 0}

    new_version = state[key]["version_count"] + 1
    img_base = _images_dir(cid)
    brief_dir = img_base / f"{num:02d}"
    brief_dir.mkdir(exist_ok=True)

    img_path = brief_dir / f"v{new_version:03d}.jpg"
    img_path.write_bytes(resp.content)

    state[key]["version_count"] = new_version
    save_state(state, cid)

    img_prefix = f"/images/{cid}/" if cid else "/images/"
    return {
        "ok": True,
        "num": num,
        "version": new_version,
        "path": f"{img_prefix}{num:02d}/v{new_version:03d}.jpg",
    }


@app.get("/api/versions/{num}")
async def get_versions(num: int):
    cid = _cartridge_id()
    state = load_state(cid)
    key = str(num)
    entry = state.get(key, {"qc_status": "pending", "selected_version": None, "version_count": 0})
    vc = entry["version_count"]
    sel = entry["selected_version"]

    img_base = _images_dir(cid)
    img_prefix = f"/images/{cid}/" if cid else "/images/"
    versions = []
    for v in range(1, vc + 1):
        path = f"{img_prefix}{num:02d}/v{v:03d}.jpg"
        exists = (img_base / f"{num:02d}" / f"v{v:03d}.jpg").exists()
        versions.append({"version": v, "path": path, "selected": v == sel, "exists": exists})

    return {"num": num, "versions": versions, "selected_version": sel, "version_count": vc}


@app.post("/api/select/{num}/{version}")
async def select_version(num: int, version: int):
    cid = _cartridge_id()
    state = load_state(cid)
    key = str(num)
    entry = state.get(key)
    if entry is None:
        raise HTTPException(status_code=404, detail=f"Brief #{num} not found")
    if version < 0 or version > entry["version_count"]:
        raise HTTPException(status_code=400, detail=f"Version {version} out of range")

    state[key]["selected_version"] = None if version == 0 else version
    save_state(state, cid)

    return {"ok": True, "num": num, "selected_version": state[key]["selected_version"]}


class PromptUpdate(BaseModel):
    prompt: Optional[str] = None

@app.post("/api/prompt/{num}")
async def update_prompt(num: int, body: PromptUpdate):
    cid = _cartridge_id()
    state = load_state(cid)
    key = str(num)
    if key not in state:
        raise HTTPException(status_code=404, detail=f"Brief #{num} not found")

    if body.prompt and body.prompt.strip():
        state[key]["prompt_override"] = body.prompt.strip()
        cache = load_prompt_cache()
        cache.pop(body.prompt.strip(), None)
        save_prompt_cache(cache)
    else:
        state[key].pop("prompt_override", None)

    save_state(state, cid)
    return {"ok": True, "num": num, "prompt_override": state[key].get("prompt_override")}


# ---------------------------------------------------------------------------
# Copy text editing + correction log
# ---------------------------------------------------------------------------

EDITABLE_COPY_FIELDS = ["copy_hl", "copy_body", "copy_cta"]


def _correction_log_path(cid: str = "") -> Path:
    cid = cid or _cartridge_id()
    return STATE_DIR / f"{cid}_corrections.json"


def load_corrections(cid: str = "") -> list[dict]:
    p = _correction_log_path(cid)
    if p.exists():
        with open(p, encoding="utf-8") as f:
            return json.load(f)
    return []


def save_corrections(logs: list[dict], cid: str = "") -> None:
    p = _correction_log_path(cid)
    with open(p, "w", encoding="utf-8") as f:
        json.dump(logs, f, ensure_ascii=False, indent=2)


def append_correction(num: int, field: str, original: str, corrected: str,
                      reason: str = "", cid: str = "") -> dict:
    entry = {
        "num": num,
        "field": field,
        "original": original,
        "corrected": corrected,
        "reason": reason,
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    logs = load_corrections(cid)
    logs.append(entry)
    save_corrections(logs, cid)
    return entry


class CopyUpdate(BaseModel):
    copy_hl: Optional[str] = None
    copy_body: Optional[str] = None
    copy_cta: Optional[str] = None
    reason: Optional[str] = None


@app.post("/api/copy/{num}")
async def update_copy(num: int, body: CopyUpdate):
    """Update copy text with override and log corrections."""
    cid = _cartridge_id()
    briefs = load_briefs(cid)
    brief = next((b for b in briefs if b["num"] == num), None)
    if brief is None:
        raise HTTPException(status_code=404, detail=f"Brief #{num} not found")

    state = load_state(cid)
    key = str(num)
    if key not in state:
        state[key] = {"qc_status": "pending", "selected_version": None, "version_count": 0}

    overrides = state[key].get("copy_overrides", {})
    corrections_made = []

    for field in EDITABLE_COPY_FIELDS:
        new_val = getattr(body, field, None)
        if new_val is None:
            continue
        new_val = new_val.strip()
        original = brief.get(field, "")

        if new_val == original or new_val == "":
            # Reset to original
            overrides.pop(field, None)
        else:
            overrides[field] = new_val
            # Log the correction
            correction = append_correction(
                num=num, field=field,
                original=original, corrected=new_val,
                reason=body.reason or "", cid=cid,
            )
            corrections_made.append(correction)

    if overrides:
        state[key]["copy_overrides"] = overrides
    else:
        state[key].pop("copy_overrides", None)

    save_state(state, cid)

    return {
        "ok": True,
        "num": num,
        "copy_overrides": overrides or None,
        "corrections_logged": len(corrections_made),
    }


@app.get("/api/corrections")
async def get_corrections():
    """Return correction log for current cartridge."""
    cid = _cartridge_id()
    logs = load_corrections(cid)
    return {"cartridge_id": cid, "corrections": logs, "total": len(logs)}


@app.get("/api/corrections/summary")
async def get_corrections_summary():
    """Aggregate corrections across all cartridges for regulation review."""
    summaries = []
    for p in sorted(STATE_DIR.glob("*_corrections.json")):
        with open(p, encoding="utf-8") as f:
            logs = json.load(f)
        cid = p.stem.replace("_corrections", "")
        by_field: dict[str, list] = {}
        for log in logs:
            by_field.setdefault(log["field"], []).append(log)
        summaries.append({
            "cartridge_id": cid,
            "total": len(logs),
            "by_field": {k: len(v) for k, v in by_field.items()},
            "corrections": logs,
        })
    return {"summaries": summaries}


class StatusUpdate(BaseModel):
    status: str


@app.post("/api/status/approve-all")
async def approve_all_status():
    """Set qc_status=approved for every brief in the current cartridge."""
    cid = _cartridge_id()
    briefs = load_briefs(cid)
    if not briefs:
        raise HTTPException(status_code=400, detail="No briefs loaded")

    state = load_state(cid)
    for b in briefs:
        key = str(b["num"])
        if key not in state:
            state[key] = {
                "qc_status": "pending",
                "selected_version": None,
                "version_count": 0,
            }
        state[key]["qc_status"] = "approved"
    save_state(state, cid)
    return {"ok": True, "count": len(briefs), "cartridge_id": cid or None}


@app.post("/api/status/pending-all")
async def pending_all_status():
    """Reset qc_status=pending for every brief in the current cartridge."""
    cid = _cartridge_id()
    briefs = load_briefs(cid)
    if not briefs:
        raise HTTPException(status_code=400, detail="No briefs loaded")

    state = load_state(cid)
    for b in briefs:
        key = str(b["num"])
        if key not in state:
            state[key] = {
                "qc_status": "pending",
                "selected_version": None,
                "version_count": 0,
            }
        state[key]["qc_status"] = "pending"
    save_state(state, cid)
    return {"ok": True, "count": len(briefs), "cartridge_id": cid or None}


@app.post("/api/status/{num}")
async def update_status(num: int, body: StatusUpdate):
    if body.status not in ("pending", "approved", "rejected"):
        raise HTTPException(status_code=400, detail="Status must be pending, approved, or rejected")

    cid = _cartridge_id()
    state = load_state(cid)
    key = str(num)
    if key not in state:
        raise HTTPException(status_code=404, detail=f"Brief #{num} not found")

    state[key]["qc_status"] = body.status
    save_state(state, cid)

    return {"ok": True, "num": num, "qc_status": body.status}


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

DELIVERABLES_DIR = APP_DIR / "output" / "deliverables"


@app.post("/api/export")
async def export_approved():
    cid = _cartridge_id()
    briefs = load_briefs(cid)
    state = load_state(cid)
    xlsx_name = get_cartridge_filename(cid) or "unknown"

    approved = []
    for b in briefs:
        key = str(b["num"])
        entry = state.get(key, {})
        if entry.get("qc_status") != "approved":
            continue
        vc = entry.get("version_count", 0)
        if vc == 0:
            continue
        sel = entry.get("selected_version")
        selected_v = sel if sel is not None else vc
        # Apply copy overrides so handoff uses corrected text
        display = {**b}
        for field in ["copy_hl", "copy_body", "copy_cta"]:
            override_val = entry.get("copy_overrides", {}).get(field)
            if override_val:
                display[field] = override_val
        approved.append({**display, "selected_version": selected_v,
                         "version_count": vc, "entry": entry})

    if not approved:
        raise HTTPException(status_code=400, detail="No approved briefs with images to export")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    # Folder name: C001_20260324_1800
    export_name = f"{cid}_{timestamp}"
    export_dir = DELIVERABLES_DIR / export_name
    export_dir.mkdir(parents=True, exist_ok=True)

    img_base = _images_dir(cid)
    exported = []
    total_images = 0

    for a in approved:
        num = a["num"]
        vc = a["version_count"]
        selected_v = a["selected_version"]

        # Subfolder: just the brief number (01, 02, ...)
        folder_name = f"{num:02d}"
        brief_dir = export_dir / folder_name
        brief_dir.mkdir(exist_ok=True)

        # Copy ALL versions
        src_dir = img_base / f"{num:02d}"
        copied_versions = []
        for v in range(1, vc + 1):
            src = src_dir / f"v{v:03d}.jpg"
            if not src.exists():
                continue
            is_selected = (v == selected_v)
            dst_name = f"v{v:03d}_OK.jpg" if is_selected else f"v{v:03d}.jpg"
            shutil.copy2(src, brief_dir / dst_name)
            copied_versions.append({
                "version": v,
                "filename": dst_name,
                "is_selected": is_selected,
            })
            total_images += 1

        if copied_versions:
            exported.append({
                **a,
                "folder_name": folder_name,
                "versions": copied_versions,
                "selected_filename": next(
                    (cv["filename"] for cv in copied_versions if cv["is_selected"]),
                    copied_versions[-1]["filename"],
                ),
            })

    # --- Generate handoff.html ---
    cards_html = ""
    for e in exported:
        hl = (e.get("copy_hl") or "").replace("\n", "<br>")
        body_text = (e.get("copy_body") or "").replace("\n", "<br>")
        folder = e["folder_name"]
        sel_file = e["selected_filename"]

        # Alternative thumbnails (all versions)
        alt_html = ""
        if len(e["versions"]) > 1:
            thumbs = ""
            for cv in e["versions"]:
                border = "3px solid #6d28d9" if cv["is_selected"] else "1px solid #ddd"
                label_extra = ' <span style="color:#6d28d9;font-weight:700">OK</span>' if cv["is_selected"] else ""
                thumbs += f"""
          <div style="text-align:center">
            <img src="{folder}/{cv['filename']}" style="width:100%;aspect-ratio:9/16;object-fit:cover;border-radius:6px;border:{border}">
            <div style="font-size:10px;color:#888;margin-top:4px">v{cv['version']:03d}{label_extra}</div>
          </div>"""
            alt_html = f"""
        <div style="margin-top:12px;padding-top:12px;border-top:1px solid #eee">
          <div style="font-size:11px;color:#888;margin-bottom:8px">All versions ({len(e['versions'])})</div>
          <div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(80px,1fr));gap:8px">
            {thumbs}
          </div>
        </div>"""

        note_text = (e.get("note") or "").strip()
        note_html = f'<div class="card-note"><strong>注釈:</strong> {note_text}</div>' if note_text else ""

        cards_html += f"""
    <div class="card">
      <div style="position:relative">
        <img src="{folder}/{sel_file}" alt="#{e['num']:02d}">
        <div style="position:absolute;top:8px;left:8px;background:#6d28d9;color:#fff;font-size:10px;font-weight:700;padding:3px 8px;border-radius:6px">OK TAKE</div>
        <div style="position:absolute;top:8px;right:8px;background:rgba(0,0,0,0.6);color:#fff;font-size:10px;padding:3px 8px;border-radius:6px">{len(e['versions'])} versions</div>
      </div>
      <div class="info">
        <div class="num">#{e['num']:02d} — {e.get('segment', '')}</div>
        <div class="hl">{hl}</div>
        <div class="body">{body_text}</div>
        <div class="cta">CTA: {e.get('copy_cta', '')}</div>
        <div class="tone">{e.get('tone', '')}</div>
        {note_html}
        {alt_html}
      </div>
    </div>"""

    handoff_html = f"""<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{cid} — Designer Handoff ({timestamp})</title>
<style>
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{ font-family: -apple-system, 'Hiragino Sans', sans-serif; background: #fafafa; color: #333; padding: 24px; }}
  h1 {{ font-size: 20px; margin-bottom: 4px; }}
  .source {{ font-size: 12px; color: #6d28d9; margin-bottom: 4px; }}
  .meta {{ font-size: 13px; color: #888; margin-bottom: 24px; }}
  .grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(320px, 1fr)); gap: 20px; }}
  .card {{ background: #fff; border-radius: 12px; overflow: hidden; box-shadow: 0 2px 8px rgba(0,0,0,0.08); }}
  .card > div:first-child img {{ width: 100%; aspect-ratio: 9/16; object-fit: cover; display: block; }}
  .info {{ padding: 16px; }}
  .num {{ font-size: 12px; color: #6d28d9; font-weight: 700; margin-bottom: 8px; }}
  .hl {{ font-size: 15px; font-weight: 700; line-height: 1.5; margin-bottom: 6px; }}
  .body {{ font-size: 13px; color: #666; line-height: 1.5; margin-bottom: 6px; }}
  .cta {{ font-size: 12px; color: #6d28d9; font-weight: 600; margin-bottom: 4px; }}
  .tone {{ font-size: 11px; color: #999; }}
  .card-note {{ margin-top: 8px; padding: 8px 10px; background: #fffbeb; border-left: 3px solid #f59e0b; border-radius: 0 6px 6px 0; font-size: 12px; color: #92400e; line-height: 1.5; }}
  .notes {{ margin-top: 24px; padding: 16px; background: #f0f0f0; border-radius: 8px; font-size: 13px; color: #666; line-height: 1.6; }}
</style>
</head>
<body>
<h1>{cid} — Designer Handoff</h1>
<div class="source">Source: {xlsx_name}</div>
<div class="meta">Exported: {timestamp} | {len(exported)} creatives | {total_images} total images</div>
<div class="grid">
{cards_html}
</div>
<div class="notes">
  <strong>Notes for Designer:</strong><br>
  - <strong>OK TAKE</strong> = selected version (file name ends with <code>_OK.jpg</code>)<br>
  - Other versions are alternatives — use if a different direction works better<br>
  - Image size: 1080x1920px (9:16)<br>
  - No text in images — copy is overlaid in design phase<br>
  - Colors/tone noted per creative
</div>
</body>
</html>"""

    with open(export_dir / "handoff.html", "w", encoding="utf-8") as f:
        f.write(handoff_html)

    return {
        "ok": True,
        "cartridge_id": cid,
        "source_xlsx": xlsx_name,
        "count": len(exported),
        "total_images": total_images,
        "path": str(export_dir),
        "handoff": f"output/deliverables/{export_name}/handoff.html",
    }


# ---------------------------------------------------------------------------
# Write overrides back to Excel
# ---------------------------------------------------------------------------

from openpyxl import load_workbook as _load_wb_write  # writable mode


# Reverse map: briefs.json key → possible xlsx headers
_KEY_TO_HEADERS = {
    "copy_hl": ["コピー HL（メインコピー）", "コピーHL（メインコピー）", "コピーHL"],
    "copy_body": ["コピー Body（サブテキスト）", "コピーBody（サブテキスト）", "コピーBody"],
    "copy_cta": ["CTA"],
}


@app.post("/api/cartridges/sync-xlsx")
async def sync_overrides_to_xlsx():
    """Write copy overrides back into the source xlsx, creating a new version."""
    cid = _cartridge_id()
    if not cid:
        raise HTTPException(status_code=400, detail="No cartridge loaded")

    xlsx_name = get_cartridge_filename(cid)
    if not xlsx_name:
        raise HTTPException(status_code=404, detail="Source xlsx not found in registry")

    # Resolve xlsx path
    src_path = BRIEFS_DIR / xlsx_name
    if not src_path.exists():
        src_path = APP_DIR / xlsx_name
    if not src_path.exists():
        raise HTTPException(status_code=404, detail=f"File not found: {xlsx_name}")

    state = load_state(cid)
    briefs = load_briefs(cid)

    # Collect overrides: {num: {field: value}}
    overrides_by_num: dict[int, dict[str, str]] = {}
    for b in briefs:
        key = str(b["num"])
        entry = state.get(key, {})
        co = entry.get("copy_overrides", {})
        if co:
            overrides_by_num[b["num"]] = co

    if not overrides_by_num:
        return {"ok": True, "updated": 0, "message": "No overrides to sync"}

    # Open xlsx in write mode
    wb = _load_wb_write(src_path)
    ws = None
    for name in wb.sheetnames:
        if "ブリーフ" in name or "brief" in name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    # Build header → column index map
    header_row = [str(c.value).strip() if c.value else "" for c in ws[1]]
    col_map: dict[str, int] = {}  # briefs key → column index (0-based)
    for key, candidates in _KEY_TO_HEADERS.items():
        for h_idx, h in enumerate(header_row):
            for cand in candidates:
                if cand in h:
                    col_map[key] = h_idx
                    break
            if key in col_map:
                break

    # Find No. column
    no_col = None
    for i, h in enumerate(header_row):
        if h in ("No.", "num", "No"):
            no_col = i
            break

    updated = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        # Get brief number
        if no_col is not None:
            try:
                num = int(row[no_col].value)
            except (ValueError, TypeError):
                num = row_idx - 1
        else:
            num = row_idx - 1

        if num not in overrides_by_num:
            continue

        for field, new_val in overrides_by_num[num].items():
            if field in col_map:
                cell = row[col_map[field]]
                cell.value = new_val
                updated += 1

    # Save as new file: original_synced.xlsx
    stem = Path(xlsx_name).stem
    synced_name = f"{stem}_synced.xlsx"
    synced_path = BRIEFS_DIR / synced_name
    wb.save(synced_path)
    wb.close()

    return {
        "ok": True,
        "updated": updated,
        "overrides_count": len(overrides_by_num),
        "output_file": synced_name,
        "message": f"{updated} cells updated → {synced_name}",
    }


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("app:app", host="0.0.0.0", port=8000, reload=True)

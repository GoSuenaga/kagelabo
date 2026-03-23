"""
設計図エクセル v2 — SE割り当て + BGM情報を含む
"""

import os, random, gspread
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv

load_dotenv(".env")

# === スプレッドシートからNo.1データ取得 ===
gc = gspread.oauth(credentials_filename='oauth_credentials.json', authorized_user_filename='token.json')
sh = gc.open_by_key('1gVGbo_fKC7sQ_B06P4M15VWn9XTMKOyKuv_Fi0VfJlc')
data = sh.sheet1.get_all_values()

cuts = []
current_no = ''
current_school = ''
current_course = ''
for row in data[1:]:
    if row[0]:
        current_no = row[0]
        current_school = row[1]
        current_course = row[2]
    if current_no != '1' or not row[4]:
        continue
    cuts.append({
        'num': row[4],
        'type': row[5],
        'narration': row[6],
        'telop': row[7],
        'logo': row[8],
        'logo_path': row[9],
        'en_prompt': row[11],
    })

school = current_school
out_dir = "output/workflow_002/no01"

# ============================================================
# SE / BGM 割り当て
# ============================================================
CUT_SE_MAP = {
    '1':  '04_tiktok',
    '2':  '02_negative',
    '3':  '02_negative',
    '4':  '01_impact',
    '5':  '01_impact',
    '6':  '03_neutral',
    '7':  '03_neutral',
    '8':  '03_neutral',
    '9':  '01_impact',
    '10': '01_impact',
    '11': '04_tiktok',
}

SE_CAT_LABEL = {
    '01_impact': 'インパクト',
    '02_negative': 'おとなしめ',
    '03_neutral': '普通',
    '04_tiktok': 'TikTok',
}

# SE選択（連続重複回避）
se_base = "clients/vantan/se/真面目バージョン"
se_files = {}
for cat in ['01_impact', '02_negative', '03_neutral', '04_tiktok']:
    cat_dir = f"{se_base}/{cat}"
    if os.path.exists(cat_dir):
        se_files[cat] = sorted([f for f in os.listdir(cat_dir) if f.endswith('.mp3')])
    else:
        se_files[cat] = []

se_assignments = []
prev_file = None
for cut in cuts:
    cat = CUT_SE_MAP.get(cut['num'], '03_neutral')
    candidates = se_files.get(cat, [])
    available = [f for f in candidates if f != prev_file] or candidates
    chosen = random.choice(available) if available else None
    se_assignments.append((cat, chosen))
    prev_file = chosen

# BGM
bgm_base = "clients/vantan/bgm"
bgm_moods = {}
for mood_dir in sorted(os.listdir(bgm_base)):
    mood_path = f"{bgm_base}/{mood_dir}"
    if os.path.isdir(mood_path):
        files = sorted([f for f in os.listdir(mood_path) if f.endswith('.mp3')])
        bgm_moods[mood_dir] = files

# ============================================================
# Excel書き出し
# ============================================================
wb = Workbook()

# --- 共通スタイル ---
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(name="Noto Sans JP", bold=True, color="FFFFFF", size=11)
body_font = Font(name="Noto Sans JP", size=10)
bold_font = Font(name="Noto Sans JP", bold=True, size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
cat_fills = {
    '01_impact': PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid"),
    '02_negative': PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid"),
    '03_neutral': PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    '04_tiktok': PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid"),
}
bgm_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")


def style_header(ws, row, cols):
    for col, h in enumerate(cols, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border


def style_cell(ws, row, col, value, font=None, fill=None, wrap=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font or body_font
    cell.border = thin_border
    cell.alignment = Alignment(vertical='center', wrap_text=wrap)
    if fill:
        cell.fill = fill
    return cell


# ============================================================
# Sheet 1: カット設計図
# ============================================================
ws1 = wb.active
ws1.title = "カット設計図"

headers = ["カット#", "タイプ", "ナレーション", "ナレ音量", "テロップ", "ロゴ",
           "SEカテゴリ", "SE特性", "SEファイル", "SE音量", "BGM", "BGM音量", "映像プロンプト(EN)"]
style_header(ws1, 1, headers)

vol_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

for i, (cut, (cat, se_file)) in enumerate(zip(cuts, se_assignments)):
    row = i + 2
    style_cell(ws1, row, 1, f"カット{cut['num'].zfill(2)}")
    style_cell(ws1, row, 2, cut['type'])
    style_cell(ws1, row, 3, cut['narration'], wrap=True)
    style_cell(ws1, row, 4, "100%", fill=vol_fill)
    style_cell(ws1, row, 5, cut['telop'] or "—", wrap=True)
    style_cell(ws1, row, 6, "○" if cut['logo'] == '○' else "—")
    style_cell(ws1, row, 7, cat, fill=cat_fills.get(cat))
    style_cell(ws1, row, 8, SE_CAT_LABEL.get(cat, ''))
    style_cell(ws1, row, 9, se_file or "—")
    style_cell(ws1, row, 10, "30%", fill=vol_fill)
    # BGM
    if i == 0:
        style_cell(ws1, row, 11, "01_hopeful (全カット共通)", fill=bgm_fill, wrap=True)
    else:
        style_cell(ws1, row, 11, "↑", fill=bgm_fill)
    style_cell(ws1, row, 12, "30%", fill=vol_fill)
    style_cell(ws1, row, 13, cut['en_prompt'][:120] + "..." if len(cut['en_prompt']) > 120 else cut['en_prompt'], wrap=True)

col_widths = [12, 8, 28, 8, 22, 6, 14, 12, 32, 8, 26, 8, 55]
for i, w in enumerate(col_widths, 1):
    col_letter = chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)
    ws1.column_dimensions[col_letter].width = w

# ============================================================
# Sheet 2: SE一覧
# ============================================================
ws2 = wb.create_sheet("SE一覧")
style_header(ws2, 1, ["カテゴリ", "特性", "番号", "ファイル名"])

row = 2
for cat in ['01_impact', '02_negative', '03_neutral', '04_tiktok']:
    for f in se_files.get(cat, []):
        style_cell(ws2, row, 1, cat, fill=cat_fills.get(cat))
        style_cell(ws2, row, 2, SE_CAT_LABEL.get(cat, ''))
        # 番号抽出
        se_num = f.split('_')[0] if '_' in f else f
        style_cell(ws2, row, 3, se_num)
        style_cell(ws2, row, 4, f)
        row += 1

ws2.column_dimensions['A'].width = 16
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 8
ws2.column_dimensions['D'].width = 40

# ============================================================
# Sheet 3: BGM一覧
# ============================================================
ws3 = wb.create_sheet("BGM一覧")
style_header(ws3, 1, ["Mood", "ファイル名", "キー", "テンポ", "特徴"])

bgm_details = {
    '01_hopeful': ('C major', '70 BPM', '静か→温かいコード→希望のクライマックス→優しく着地'),
    '02_tender': ('G major', '65 BPM', 'アルペジオ→メロディ→盛り上がり→シンプルに戻る'),
    '03_reflective': ('F major', '60 BPM', 'ミニマル、2音から→Maj7の明るさ→希望の一音で締め'),
}

row = 2
for mood, files in bgm_moods.items():
    detail = bgm_details.get(mood, ('', '', ''))
    for f in files:
        style_cell(ws3, row, 1, mood, fill=bgm_fill)
        style_cell(ws3, row, 2, f)
        style_cell(ws3, row, 3, detail[0])
        style_cell(ws3, row, 4, detail[1])
        style_cell(ws3, row, 5, detail[2], wrap=True)
        row += 1

ws3.column_dimensions['A'].width = 18
ws3.column_dimensions['B'].width = 35
ws3.column_dimensions['C'].width = 12
ws3.column_dimensions['D'].width = 10
ws3.column_dimensions['E'].width = 50

# ============================================================
# Sheet 4: 音量設定（マスター）
# ============================================================
ws4 = wb.create_sheet("音量設定")
style_header(ws4, 1, ["項目", "音量", "備考"])

audio_settings = [
    ("ナレーション", "100%", "主役。常にフル音量"),
    ("SE（効果音）", "30%", "カット感情に応じた4カテゴリから選択。連続同一SE禁止"),
    ("BGM", "30%", "ピアノソロ。動画全体に薄くかける。ナレーションの邪魔をしない"),
]

for i, (item, vol, note) in enumerate(audio_settings):
    row = i + 2
    style_cell(ws4, row, 1, item, font=bold_font)
    style_cell(ws4, row, 2, vol, fill=vol_fill)
    style_cell(ws4, row, 3, note, wrap=True)

# ルール記載
rule_row = len(audio_settings) + 3
style_cell(ws4, rule_row, 1, "■ 音量設計ルール", font=bold_font)
rules = [
    "ナレーションが最も重要。SE・BGMはナレーションを邪魔しない音量に抑える",
    "BGMはピアノソロ、メジャーキー、マイナー禁止",
    "SEは同じファイルが連続カットで使われないようにする",
    "BGMのMood: 01_hopeful / 02_tender / 03_reflective から動画に合うものを選択",
]
for i, rule in enumerate(rules):
    style_cell(ws4, rule_row + 1 + i, 1, f"  {i+1}. {rule}", wrap=True)
    ws4.merge_cells(start_row=rule_row + 1 + i, start_column=1, end_row=rule_row + 1 + i, end_column=3)

ws4.column_dimensions['A'].width = 20
ws4.column_dimensions['B'].width = 10
ws4.column_dimensions['C'].width = 60

# 保存
excel_path = f"{out_dir}/no01_設計図.xlsx"
wb.save(excel_path)
print(f"✓ {excel_path}")

# サマリー表示
print(f"\nシート1: カット設計図（{len(cuts)}カット × SE・BGM割り当て）")
print(f"シート2: SE一覧（{sum(len(v) for v in se_files.values())}個）")
print(f"シート3: BGM一覧（{sum(len(v) for v in bgm_moods.values())}曲）")

print(f"\nBGM試聴ファイル:")
for mood, files in bgm_moods.items():
    detail = bgm_details.get(mood, ('', '', ''))
    for f in files:
        print(f"  {mood}/{f}  [{detail[0]}, {detail[1]}] {detail[2]}")

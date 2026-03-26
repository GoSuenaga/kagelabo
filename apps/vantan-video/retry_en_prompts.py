"""リトライ: 未更新の10件のENプロンプトを再生成"""
import openpyxl, os, re, time
from pathlib import Path
from dotenv import load_dotenv
from google import genai

load_dotenv(Path(__file__).parent.parent.parent / ".env")

XLSX = Path(__file__).parent / "briefs" / "VANTAN_台本_v3.0_20260326.xlsx"
OLD_XLSX = Path(__file__).parent / "briefs" / "VANTAN_台本_v2.0_20260323.xlsx"

key = os.getenv("GEMINI_API_KEY_2")
client = genai.Client(api_key=key)

old_wb = openpyxl.load_workbook(OLD_XLSX, data_only=True)
wb = openpyxl.load_workbook(XLSX)
ws_old = old_wb["台本"]
ws = wb["台本"]
headers = [c.value for c in ws[1]]
en_col = headers.index("映像プロンプト（EN）") + 1
jp_col = headers.index("映像プロンプト（日本語）") + 1

SYSTEM = (
    "You are an expert cinematographer. Convert the Japanese video prompt into an optimal English prompt for Google Veo 3.1. "
    "RULES: 1) 100% faithful to Japanese (exact age, clothing, hair, location, props, actions). "
    "2) NEVER add elements not in JP. 3) Always say 'Japanese' for people. "
    "4) Append: 'Shot on a high-end cinema camera with anamorphic lens. Documentary film aesthetic, warm cinematic color grading. No text, no logos, no signage.' "
    "5) Add ONE camera movement. 6) 100-200 words. Output ONLY English prompt."
)

fixed = 0
for row in range(2, ws.max_row + 1):
    old_en = (ws_old.cell(row, en_col).value or "").strip()
    new_en = (ws.cell(row, en_col).value or "").strip()
    jp = (ws.cell(row, jp_col).value or "").strip()
    if old_en == new_en and jp:
        no = ws.cell(row, 1).value or "?"
        cut = ws.cell(row, headers.index("カット#") + 1).value or "?"
        print(f"リトライ: no{str(no).zfill(2)} カット{cut}...", end=" ", flush=True)
        try:
            resp = client.models.generate_content(
                model="gemini-2.5-flash",
                contents=[{"role": "user", "parts": [{"text": SYSTEM + "\n\n日本語プロンプト:\n" + jp}]}],
            )
            en = resp.text.strip() if resp.text else ""
            en = re.sub(r"^```.*\n?", "", en).strip().strip('"').strip("'")
            if en and len(en) > 20:
                ws.cell(row, en_col).value = en
                fixed += 1
                print(f"OK ({len(en.split())}語)")
            else:
                print("空レスポンス")
        except Exception as e:
            print(f"ERR: {str(e)[:80]}")
        time.sleep(1)

wb.save(XLSX)
print(f"\n修復: {fixed}/10")

"""
設計図エクセル書き出し + SE重複回避ロジック付き動画再合成
"""

import requests, time, os, random, gspread
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv

load_dotenv()
fal_key = os.getenv("FAL_API_KEY")
creatomate_key = os.getenv("CREATOMATE_API_KEY")

# === スプレッドシートからNo.1データ取得 ===
gc = gspread.oauth(credentials_filename='oauth_credentials.json', authorized_user_filename='token.json')
sh = gc.open_by_key('1gVGbo_fKC7sQ_B06P4M15VWn9XTMKOyKuv_Fi0VfJlc')
data = sh.sheet1.get_all_values()

cuts = []
current_no = ''
current_school = ''
current_course = ''
for row in data[1:]:
    if row[0]:
        current_no = row[0]
        current_school = row[1]
        current_course = row[2]
    if current_no != '1' or not row[4]:
        continue
    cuts.append({
        'num': row[4],
        'type': row[5],
        'narration': row[6],
        'telop': row[7],
        'logo': row[8],
        'logo_path': row[9],
        'en_prompt': row[11],
    })

school = current_school
out_dir = "output/workflow_002/no01"

# ============================================================
# SEカテゴリ割り当て
# ============================================================
CUT_SE_MAP = {
    '1':  '04_tiktok',
    '2':  '02_negative',
    '3':  '02_negative',
    '4':  '01_impact',
    '5':  '01_impact',
    '6':  '03_neutral',
    '7':  '03_neutral',
    '8':  '03_neutral',
    '9':  '01_impact',
    '10': '01_impact',
    '11': '04_tiktok',
}

SE_CAT_LABEL = {
    '01_impact': 'インパクト',
    '02_negative': 'おとなしめ',
    '03_neutral': '普通',
    '04_tiktok': 'TikTok',
}

# ============================================================
# SE選択（連続同一SE回避）
# ============================================================
se_base = "clients/vantan/se/真面目バージョン"
se_files = {}
for cat in ['01_impact', '02_negative', '03_neutral', '04_tiktok']:
    cat_dir = f"{se_base}/{cat}"
    if os.path.exists(cat_dir):
        se_files[cat] = sorted([f for f in os.listdir(cat_dir) if f.endswith('.mp3')])
    else:
        se_files[cat] = []


def pick_se_no_repeat(cut_list):
    """連続して同じSEにならないように選択"""
    result = []
    prev_file = None
    for cut in cut_list:
        cat = CUT_SE_MAP.get(cut['num'], '03_neutral')
        candidates = se_files.get(cat, [])
        if not candidates:
            result.append((cat, None))
            prev_file = None
            continue
        # 前回と同じファイルを除外
        available = [f for f in candidates if f != prev_file]
        if not available:
            available = candidates  # 1個しかない場合はしょうがない
        chosen = random.choice(available)
        result.append((cat, chosen))
        prev_file = chosen
    return result


se_assignments = pick_se_no_repeat(cuts)

# ============================================================
# Excel書き出し
# ============================================================
print("設計図エクセル書き出し中...")

wb = Workbook()
ws = wb.active
ws.title = "No.01 設計図"

# ヘッダー
headers_row = ["カット#", "タイプ", "ナレーション", "テロップ", "ロゴ", "SEカテゴリ", "SE特性", "SEファイル", "映像プロンプト(EN)"]
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(name="Noto Sans JP", bold=True, color="FFFFFF", size=11)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)

for col, h in enumerate(headers_row, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# カテゴリ色
cat_fills = {
    '01_impact': PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid"),
    '02_negative': PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid"),
    '03_neutral': PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    '04_tiktok': PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid"),
}

body_font = Font(name="Noto Sans JP", size=10)

for i, (cut, (cat, se_file)) in enumerate(zip(cuts, se_assignments)):
    row = i + 2
    values = [
        f"カット{cut['num'].zfill(2)}",
        cut['type'],
        cut['narration'],
        cut['telop'] if cut['telop'] else "—",
        "○" if cut['logo'] == '○' else "—",
        cat,
        SE_CAT_LABEL.get(cat, ''),
        se_file or "—",
        cut['en_prompt'][:100] + "..." if len(cut['en_prompt']) > 100 else cut['en_prompt'],
    ]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = body_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=(col in [3, 4, 9]))
        # SEカテゴリ列に色付け
        if col == 6:
            cell.fill = cat_fills.get(cat, PatternFill())

# 列幅調整
col_widths = [12, 8, 30, 25, 6, 16, 14, 30, 60]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[chr(64 + i)].width = w

# SE一覧シート
ws2 = wb.create_sheet("SE一覧")
ws2.cell(row=1, column=1, value="カテゴリ").font = header_font
ws2.cell(row=1, column=1).fill = header_fill
ws2.cell(row=1, column=2, value="特性").font = header_font
ws2.cell(row=1, column=2).fill = header_fill
ws2.cell(row=1, column=3, value="ファイル名").font = header_font
ws2.cell(row=1, column=3).fill = header_fill

row = 2
for cat in ['01_impact', '02_negative', '03_neutral', '04_tiktok']:
    for f in se_files.get(cat, []):
        ws2.cell(row=row, column=1, value=cat).font = body_font
        ws2.cell(row=row, column=1).fill = cat_fills.get(cat, PatternFill())
        ws2.cell(row=row, column=2, value=SE_CAT_LABEL.get(cat, '')).font = body_font
        ws2.cell(row=row, column=3, value=f).font = body_font
        row += 1

ws2.column_dimensions['A'].width = 16
ws2.column_dimensions['B'].width = 14
ws2.column_dimensions['C'].width = 40

excel_path = f"{out_dir}/no01_設計図.xlsx"
wb.save(excel_path)
print(f"✓ {excel_path}")

# ============================================================
# 動画再合成（SE重複回避版）
# ============================================================
print(f"\n動画合成（SE重複回避）...")


def upload(filepath, content_type):
    with open(filepath, 'rb') as f:
        d = f.read()
    init = requests.post(
        "https://rest.alpha.fal.ai/storage/upload/initiate",
        headers={"Authorization": f"Key {fal_key}", "Content-Type": "application/json"},
        json={"file_name": os.path.basename(filepath), "content_type": content_type},
        timeout=30,
    )
    requests.put(init.json()["upload_url"], data=d, headers={"Content-Type": content_type}, timeout=60)
    return init.json()["file_url"]


# 動画・音声アップロード
vid_urls, audio_urls = {}, {}
for cut in cuts:
    num = cut['num']
    vp = f"{out_dir}/videos/カット{num.zfill(2)}.mp4"
    ap = f"{out_dir}/audio/カット{num.zfill(2)}.mp3"
    if os.path.exists(vp):
        vid_urls[num] = upload(vp, "video/mp4")
    if os.path.exists(ap):
        audio_urls[num] = upload(ap, "audio/mpeg")

# ロゴ
logo_url = ""
for c in cuts:
    if c['logo'] == '○' and c['logo_path'] and os.path.exists(c['logo_path']):
        logo_url = upload(c['logo_path'], "image/jpeg")
        break

# SE: 割り当て済みファイルをアップロード
se_urls = []
for cat, se_file in se_assignments:
    if se_file:
        url = upload(f"{se_base}/{cat}/{se_file}", "audio/mpeg")
        se_urls.append(url)
        print(f"  SE: {se_file}")
    else:
        se_urls.append("")

# Creatomate elements
elements = []
for i, cut in enumerate(cuts):
    num = cut['num']
    if num not in vid_urls or num not in audio_urls:
        continue

    scene_elements = [
        {"type": "video", "source": vid_urls[num], "fit": "cover", "duration": "100%", "loop": True},
        {"type": "audio", "source": audio_urls[num]},
    ]

    if cut['logo'] == '○' and logo_url:
        scene_elements.append({
            "type": "image", "source": logo_url,
            "x": "50%", "y": "50%", "width": "75%", "height": "25%",
            "fit": "contain", "x_alignment": "50%", "y_alignment": "50%", "z_index": 15,
        })
    elif cut['telop']:
        scene_elements.append({
            "type": "text", "text": cut['telop'],
            "width": "85%", "height": "20%", "x": "50%", "y": "50%",
            "duration": "100%", "z_index": 15,
            "fill_color": "#FFFFFF", "font_family": "Noto Sans JP", "font_weight": "900",
            "shadow_color": "rgba(0,0,0,0.6)", "shadow_blur": "25px",
            "x_alignment": "50%", "y_alignment": "50%", "content_alignment": "center",
            "dynamic_font_size": True, "font_size_maximum": "70px", "font_size_minimum": "30px",
            "fit": "shrink",
        })

    if se_urls[i]:
        scene_elements.append({"type": "audio", "source": se_urls[i], "duration": "100%"})

    scene = {"type": "composition", "track": 1, "elements": scene_elements}
    if len(elements) > 0:
        scene["transition"] = {"type": "crossfade", "duration": 0.1}
    elements.append(scene)

print(f"\n合成カット数: {len(elements)}")

cr_resp = requests.post(
    "https://api.creatomate.com/v1/renders",
    headers={"Authorization": f"Bearer {creatomate_key}", "Content-Type": "application/json"},
    json={"source": {"output_format": "mp4", "frame_rate": 30, "width": 720, "height": 1280, "elements": elements}},
    timeout=60,
)
renders = cr_resp.json()
render_obj = renders[0] if isinstance(renders, list) else renders
render_id = render_obj.get("id", "")
print(f"レンダリング開始: {render_id}")

for i in range(60):
    time.sleep(10)
    poll = requests.get(
        f"https://api.creatomate.com/v1/renders/{render_id}",
        headers={"Authorization": f"Bearer {creatomate_key}"},
        timeout=30,
    )
    status = poll.json().get("status", "")
    print(f"  ポーリング {i+1}: {status}")
    if status == "succeeded":
        final_url = poll.json().get("url", "")
        vid = requests.get(final_url, timeout=120)
        final_path = f"{out_dir}/final.mp4"
        with open(final_path, 'wb') as f:
            f.write(vid.content)
        print(f"\n{'='*60}")
        print(f"✓ 完成: {final_path} ({len(vid.content)//1024}KB)")
        print(f"✓ 設計図: {excel_path}")
        print(f"{'='*60}")
        break
    elif status == "failed":
        print(f"  ✗ 合成失敗: {poll.json().get('error_message', '')}")
        break
else:
    print("  ✗ タイムアウト")

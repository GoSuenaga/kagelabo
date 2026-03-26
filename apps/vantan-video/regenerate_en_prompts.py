"""
ENプロンプト再生成 — JPプロンプトを正として Veo 3.1 最適化ENを生成
Usage: cd apps/vantan-video && python3 regenerate_en_prompts.py

出力: briefs/VANTAN_台本_v3.0_YYYYMMDD.xlsx
"""

import os, time, re, copy
from datetime import datetime
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from google import genai

load_dotenv(Path(__file__).parent.parent.parent / ".env")

# ── 設定 ──
INPUT_XLSX = Path(__file__).parent / "briefs" / "VANTAN_台本_v2.0_20260323.xlsx"
TODAY = datetime.now().strftime("%Y%m%d")
OUTPUT_XLSX = Path(__file__).parent / "briefs" / f"VANTAN_台本_v3.0_{TODAY}.xlsx"

# Gemini キー（テキスト生成用 — Veo クォータとは別）
GEMINI_KEYS = []
for i in range(1, 10):
    k = os.getenv(f"GEMINI_API_KEY_{i}", "")
    if k:
        GEMINI_KEYS.append(k)
if not GEMINI_KEYS:
    raise RuntimeError("GEMINI_API_KEY_1 が .env にありません")

_key_idx = 0


def get_client():
    return genai.Client(api_key=GEMINI_KEYS[_key_idx])


def rotate_key():
    global _key_idx
    _key_idx = (_key_idx + 1) % len(GEMINI_KEYS)
    print(f"  → キー切替: KEY_{_key_idx + 1}")


# ── Veo 3.1 最適化翻訳プロンプト ──
SYSTEM_PROMPT = """You are an expert cinematographer and video prompt engineer specializing in Google Veo 3.1 video generation.

Your task: Convert a Japanese video prompt into an optimal English prompt for Veo 3.1.

## ABSOLUTE RULES — NEVER VIOLATE THESE

1. **100% faithful to the Japanese source**: Do NOT change, add, or remove ANY information from the Japanese prompt.
   - Age: translate exactly (15歳 → 15-year-old, NOT 16)
   - Clothing: translate exactly (白Tシャツにデニム → white T-shirt and denim jeans, NOT henley shirt)
   - Hair: translate exactly (短めの黒髪 → short black hair, NOT dark brown)
   - Location: translate exactly — do NOT add tatami, shoji, engawa, or any traditional Japanese elements unless explicitly mentioned in the Japanese
   - Props: translate exactly (色鉛筆やマーカー → colored pencils and markers, NOT tea cups)
   - Actions: translate exactly (窓辺に立つ → standing by the window, NOT kneeling)
   - Mood/atmosphere: translate exactly
   - Time of day: translate exactly (午後 → afternoon, NOT evening)

2. **NEVER add elements not in the Japanese prompt**: No invented furniture, decorations, props, or setting details.

3. **Always specify "Japanese" for people**: Every person must be described as Japanese (e.g., "15-year-old Japanese boy").

## WHAT YOU SHOULD ADD (these are the ONLY additions allowed)

After faithfully translating all content from the Japanese prompt, append these cinematic specifications:

1. **Camera work** (pick ONE that fits the scene):
   - Static shots: "Static shot" or "Slow dolly in"
   - Movement scenes: "Slow tracking shot" or "Follow shot"
   - Wide/landscape: "Slow pan" or "Crane shot"
   - Close-ups: "Shallow depth of field, natural bokeh"

2. **Cinematic style suffix** (always append):
   "Shot on a high-end cinema camera with anamorphic lens. Documentary film aesthetic, warm cinematic color grading. No text, no logos, no signage."

3. **Subtle movement instruction** if the Japanese prompt describes a static scene:
   Add one small motion element (e.g., "hair gently moving in the breeze", "light slowly shifting")

## OUTPUT FORMAT

- English only
- 100-200 words (3-6 sentences)
- Structure: [Camera/Shot type] → [Setting] → [Subject & appearance] → [Action] → [Cinematic style]
- Do NOT use negative instructions like "no walls" — use noun-list format if needed
- Do NOT include any explanation — output ONLY the English prompt

## IMPORTANT CONTEXT

These prompts are for school advertisement videos (VANTAN schools). The style is:
- Documentary/cinematic film aesthetic (NOT iPhone/social media style)
- Shallow depth of field, bokeh, natural lighting
- Emotional storytelling with warm color grading
"""

USER_TEMPLATE = """以下の日本語映像プロンプトを、Veo 3.1 用の最適な英語プロンプトに変換してください。

## 情報
- スクール名: {school}
- 分野: {field}
- 対象の子ども: {child}
- カット#: {cut_num}
- カットタイプ: {cut_type}
- ナレーション: {narration}
- 動画のムード: {mood}

## 日本語映像プロンプト（これが正。100%忠実に英語化すること）
{prompt_jp}
"""


def translate_prompt(jp_prompt, school, field, child, cut_num, cut_type, narration, mood):
    """JP → Veo 3.1 最適化 EN"""
    if not jp_prompt or not jp_prompt.strip():
        return ""

    user_msg = USER_TEMPLATE.format(
        school=school or "",
        field=field or "",
        child=child or "",
        cut_num=cut_num or "",
        cut_type=cut_type or "",
        narration=narration or "",
        mood=mood or "",
        prompt_jp=jp_prompt,
    )

    for attempt in range(len(GEMINI_KEYS)):
        try:
            client = get_client()
            resp = client.models.generate_content(
                model="gemini-2.5-flash",
                contents=[
                    {"role": "user", "parts": [{"text": SYSTEM_PROMPT + "\n\n" + user_msg}]},
                ],
            )
            en = resp.text.strip()
            # マークダウンの装飾を除去
            en = re.sub(r'^```.*\n?', '', en)
            en = re.sub(r'```$', '', en)
            en = en.strip().strip('"').strip("'")
            return en
        except Exception as e:
            err = str(e)
            if "quota" in err.lower() or "429" in err or "resource_exhausted" in err.lower():
                rotate_key()
                continue
            print(f"  ⚠ エラー: {err[:100]}")
            return f"ERROR: {err[:200]}"

    return "ERROR: 全キーがクォータ超過"


def main():
    print(f"📂 入力: {INPUT_XLSX}")
    print(f"📂 出力: {OUTPUT_XLSX}")
    print(f"🔑 Gemini キー: {len(GEMINI_KEYS)}本")
    print()

    wb = openpyxl.load_workbook(INPUT_XLSX)
    ws = wb["台本"]
    headers = [c.value for c in ws[1]]

    col = {h: i + 1 for i, h in enumerate(headers)}

    total = 0
    translated = 0
    errors = 0

    for row in range(2, ws.max_row + 1):
        jp = ws.cell(row, col["映像プロンプト（日本語）"]).value
        if not jp or not jp.strip():
            continue

        total += 1
        no = ws.cell(row, col["No"]).value or ""
        cut_num = ws.cell(row, col["カット#"]).value or ""
        school = ws.cell(row, col["スクール名称"]).value or ""
        field = ws.cell(row, col["分野"]).value or ""
        child = ws.cell(row, col["子ども"]).value or ""
        cut_type = ws.cell(row, col["カットタイプ"]).value or ""
        narration = ws.cell(row, col["ナレーション"]).value or ""
        mood = ws.cell(row, col["動画のムード"]).value or ""

        pat_label = f"no{str(no).zfill(2)}" if no else "?"
        print(f"[{translated+1}/{total}] {pat_label} カット{cut_num} ...", end=" ", flush=True)

        en = translate_prompt(jp, school, field, child, cut_num, cut_type, narration, mood)

        if en.startswith("ERROR"):
            errors += 1
            print(f"❌ {en[:80]}")
        else:
            translated += 1
            # EN列を上書き
            ws.cell(row, col["映像プロンプト（EN）"]).value = en
            # 文字数チェック
            word_count = len(en.split())
            print(f"✅ ({word_count}語)")

        # レート制限対策
        time.sleep(0.5)

    # 保存
    wb.save(OUTPUT_XLSX)
    print()
    print(f"🎉 完了！")
    print(f"   翻訳: {translated}/{total} カット")
    print(f"   エラー: {errors}")
    print(f"   出力: {OUTPUT_XLSX}")
    print()
    print("次のステップ:")
    print(f"  1. Video Studio (localhost:8888) で XLSX再読込")
    print(f"  2. カートリッジ「VANTAN_台本_v3.0_{TODAY}.xlsx」を選択")
    print(f"  3. 動画生成を再開")


if __name__ == "__main__":
    main()

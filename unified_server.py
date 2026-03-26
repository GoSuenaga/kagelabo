#!/usr/bin/env python3
"""
unified_server.py — kage-lab Unified Mobile Studio v2.0
CCM + VANTAN Video Studio を統合。Cloudflare Tunnel で5G/外出先からもアクセス可能。

使い方: python3 unified_server.py
→ ローカル: http://localhost:8888
→ 公開URL: ターミナルに表示される https://xxx.trycloudflare.com
"""

from __future__ import annotations

import asyncio
import json
import glob as glob_mod
import os
import re
import secrets
import shutil
import subprocess
import threading
import time
import random
import signal
import sys
from datetime import datetime
from pathlib import Path

try:
    from fastapi import FastAPI, Request, Response, Form, BackgroundTasks, UploadFile, File
    from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
    from fastapi.staticfiles import StaticFiles
    from fastapi.middleware.cors import CORSMiddleware
    import uvicorn
except ImportError:
    print("FastAPI/uvicorn が必要です。インストールします...")
    subprocess.run(["python3", "-m", "pip", "install", "fastapi", "uvicorn[standard]"], check=True)
    from fastapi import FastAPI, Request, Response, Form, BackgroundTasks, UploadFile, File
    from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
    from fastapi.staticfiles import StaticFiles
    from fastapi.middleware.cors import CORSMiddleware
    import uvicorn

try:
    from openpyxl import load_workbook
except ImportError:
    subprocess.run(["python3", "-m", "pip", "install", "openpyxl"], check=True)
    from openpyxl import load_workbook

# ── Config ──────────────────────────────────────────────
VERSION = "2.0"
PORT = 8888
ROOT = os.path.dirname(os.path.abspath(__file__))
WORK_DIR = Path(ROOT)
VANTAN = os.path.join(ROOT, "apps", "vantan-video")
CLIENTS = os.path.join(ROOT, "clients")
BRIEFS_DIR = os.path.join(VANTAN, "briefs")
STATE_DIR = os.path.join(VANTAN, "state")
META_PATH = os.path.join(VANTAN, "cartridge_meta.json")
STATUS_MD = WORK_DIR / "apps" / "vantan-video" / "STATUS.md"
os.makedirs(BRIEFS_DIR, exist_ok=True)
os.makedirs(STATE_DIR, exist_ok=True)

# CCM Auth
PASSWORD = os.environ.get("MOBILE_CLAUDE_PW", "")
if not PASSWORD:
    PASSWORD = secrets.token_urlsafe(12)

# Claude CLI path
CLAUDE_PATH = None
for p in [
    Path.home() / ".claude" / "local" / "claude",
    Path("/opt/homebrew/bin/claude"),
    Path("/usr/local/bin/claude"),
]:
    if p.exists():
        CLAUDE_PATH = str(p)
        break

def find_claude():
    if CLAUDE_PATH:
        return CLAUDE_PATH
    try:
        result = subprocess.run(["which", "claude"], capture_output=True, text=True)
        if result.returncode == 0:
            return result.stdout.strip()
    except Exception:
        pass
    return "claude"

# ── Tunnel ──────────────────────────────────────────────
tunnel_url = ""
tunnel_proc = None

def start_tunnel():
    global tunnel_url, tunnel_proc
    try:
        tunnel_proc = subprocess.Popen(
            ["cloudflared", "tunnel", "--url", f"http://localhost:{PORT}"],
            stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True,
        )
        for line in iter(tunnel_proc.stdout.readline, ""):
            m = re.search(r"(https://[a-z0-9-]+\.trycloudflare\.com)", line)
            if m:
                tunnel_url = m.group(1)
                print(f"\n  PUBLIC URL: {tunnel_url}")
                print(f"  (スマホでこのURLを開いてください)\n")
                break
    except FileNotFoundError:
        print("  cloudflared 未インストール: brew install cloudflared")
    except Exception as e:
        print(f"  Tunnel error: {e}")

# ── State (CCM) ────────────────────────────────────────
server_start = time.time()
valid_sessions: set[str] = set()
claude_sessions: dict[str, str] = {}
claude_busy = False
message_queue: dict[str, list[dict]] = {}

# ── Launchable services ────────────────────────────────
SERVICES = {
    "rag": {"name": "RAG Creative Studio", "cmd": ["python3", "app.py"], "cwd": str(WORK_DIR / "apps" / "rag-images"), "port": 8000},
    "vlog": {"name": "Vlog 動画生成 UI", "cmd": ["python3", "-m", "streamlit", "run", "vlog_app.py"], "cwd": str(WORK_DIR / "apps" / "vantan-video"), "port": 8501},
}
launched_procs: dict[str, subprocess.Popen] = {}

# ============================================================
# カートリッジシステム（Studio から移植）
# ============================================================
REGISTRY_PATH = os.path.join(STATE_DIR, "cartridge_registry.json")

def load_registry():
    if os.path.exists(REGISTRY_PATH):
        with open(REGISTRY_PATH, encoding="utf-8") as f:
            return json.load(f)
    return {"next_id": 1, "cartridges": {}}

def save_registry(reg):
    with open(REGISTRY_PATH, "w", encoding="utf-8") as f:
        json.dump(reg, f, ensure_ascii=False, indent=2)

def register_cartridge(filename):
    reg = load_registry()
    for cid, info in reg.get("cartridges", {}).items():
        if info["filename"] == filename:
            return cid
    num = reg.get("next_id", 1)
    cid = f"C{num:03d}"
    reg.setdefault("cartridges", {})[cid] = {
        "filename": filename,
        "registered_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    reg["next_id"] = num + 1
    save_registry(reg)
    return cid

def load_meta():
    if os.path.exists(META_PATH):
        with open(META_PATH, encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_meta(meta):
    with open(META_PATH, "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)

def get_active_cid():
    return load_meta().get("cartridge_id", "")

def set_active_cid(cid, filename="", pattern_count=0):
    save_meta({
        "cartridge_id": cid, "loaded_file": filename,
        "loaded_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "pattern_count": pattern_count,
    })

def _patterns_path(cid):
    return os.path.join(STATE_DIR, f"{cid}_patterns.json")

def _state_path(cid):
    return os.path.join(STATE_DIR, f"{cid}_state.json")

def _corrections_path(cid):
    return os.path.join(STATE_DIR, f"{cid}_corrections.json")

def load_cartridge_patterns(cid):
    p = _patterns_path(cid)
    if os.path.exists(p):
        with open(p, encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_cartridge_patterns(cid, patterns):
    with open(_patterns_path(cid), "w", encoding="utf-8") as f:
        json.dump(patterns, f, ensure_ascii=False, indent=2)

def cartridge_output_base(cid):
    d = os.path.join(VANTAN, "output", cid)
    os.makedirs(d, exist_ok=True)
    return d

# ============================================================
# XLSX パーサー
# ============================================================
COLUMN_MAP = {
    "No": "no", "スクール名称": "school", "分野": "field", "子ども": "child",
    "カット#": "num", "カットタイプ": "cut_type", "ナレーション": "narration",
    "テロップ": "telop", "ロゴ表示": "logo", "ロゴファイルパス": "logo_path",
    "映像プロンプト（日本語）": "prompt_jp", "映像プロンプト（EN）": "prompt_en",
    "LP": "lp", "動画のムード": "mood",
}

def parse_xlsx_to_patterns(filepath):
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = None
    for name in wb.sheetnames:
        if "台本" in name or "brief" in name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return {}
    headers = [str(h).strip() if h else "" for h in rows[0]]
    col_idx = {}
    for i, h in enumerate(headers):
        if h in COLUMN_MAP:
            col_idx[COLUMN_MAP[h]] = i
        else:
            for pattern, key in COLUMN_MAP.items():
                if pattern in h:
                    col_idx[key] = i
                    break
    def cell(row, key):
        idx = col_idx.get(key)
        if idx is None or idx >= len(row):
            return ""
        return str(row[idx]).strip() if row[idx] is not None else ""
    patterns = {}
    current_no, current_school, current_field, current_child = "", "", "", ""
    for row in rows[1:]:
        no_val = cell(row, "no")
        if no_val and no_val != "None":
            current_no = no_val
            current_school = cell(row, "school") or current_school
            current_field = cell(row, "field") or current_field
            current_child = cell(row, "child") or current_child
        cut_num = cell(row, "num")
        if not cut_num or cut_num == "None":
            continue
        key = f"no{current_no.zfill(2)}"
        if key not in patterns:
            patterns[key] = {"school": current_school, "field": current_field, "child": current_child, "cuts": []}
        patterns[key]["cuts"].append({
            "num": cut_num, "cut_type": cell(row, "cut_type"), "narration": cell(row, "narration"),
            "telop": cell(row, "telop"), "logo": cell(row, "logo"), "logo_path": cell(row, "logo_path"),
            "prompt_jp": cell(row, "prompt_jp"), "prompt_en": cell(row, "prompt_en"),
            "lp": cell(row, "lp"), "mood": cell(row, "mood"),
        })
    return patterns

# ============================================================
# 状態管理
# ============================================================
_state_lock = threading.Lock()

def load_state():
    cid = get_active_cid()
    if not cid:
        return {}
    sf = _state_path(cid)
    with _state_lock:
        try:
            if os.path.exists(sf):
                with open(sf, encoding="utf-8") as f:
                    return json.load(f)
        except (json.JSONDecodeError, IOError):
            pass
        return {}

def save_state(state):
    cid = get_active_cid()
    if not cid:
        return
    sf = _state_path(cid)
    with _state_lock:
        tmp = sf + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(state, f, ensure_ascii=False, indent=2)
        os.replace(tmp, sf)

def get_pattern_state(pat_key):
    state = load_state()
    return state.get(pat_key, {"status": "idle", "progress": "", "message": ""})

def set_pattern_state(pat_key, **kwargs):
    state = load_state()
    if pat_key not in state:
        state[pat_key] = {"status": "idle", "progress": "", "message": "", "stop_requested": False}
    state[pat_key].update(kwargs)
    state[pat_key]["updated"] = datetime.now().strftime("%H:%M:%S")
    save_state(state)

def get_cut_state(pat_key, cut_num):
    state = load_state()
    pat = state.get(pat_key, {})
    cuts = pat.get("cuts_qc", {})
    return cuts.get(str(cut_num), {
        "qc_status": "pending", "narration_override": None,
        "telop_override": None, "prompt_jp_override": None, "prompt_en_override": None,
    })

def set_cut_state(pat_key, cut_num, **kwargs):
    state = load_state()
    if pat_key not in state:
        state[pat_key] = {"status": "idle", "progress": "", "message": "", "stop_requested": False}
    if "cuts_qc" not in state[pat_key]:
        state[pat_key]["cuts_qc"] = {}
    cn = str(cut_num)
    if cn not in state[pat_key]["cuts_qc"]:
        state[pat_key]["cuts_qc"][cn] = {"qc_status": "pending"}
    state[pat_key]["cuts_qc"][cn].update(kwargs)
    save_state(state)

# ============================================================
# アクティブパターン
# ============================================================
PATTERNS = {}

def reload_active_patterns():
    global PATTERNS
    cid = get_active_cid()
    if cid:
        PATTERNS = load_cartridge_patterns(cid)
    else:
        PATTERNS = {}
    return PATTERNS

def get_output_base():
    cid = get_active_cid()
    if cid:
        return cartridge_output_base(cid)
    return os.path.join(VANTAN, "output", "_no_cartridge")

# 起動時: アクティブカートリッジを読み込む
cid = get_active_cid()
if cid:
    print(f"  Cartridge {cid} loading...")
    PATTERNS = load_cartridge_patterns(cid)
    meta = load_meta()
    print(f"  OK: {meta.get('loaded_file', '?')} ({len(PATTERNS)} patterns)")

# ============================================================
# 出力チェック
# ============================================================
def check_outputs(pat_key):
    base = os.path.join(get_output_base(), pat_key)
    vid_dir, aud_dir = os.path.join(base, "videos"), os.path.join(base, "audio")
    videos = sorted([f for f in os.listdir(vid_dir) if f.endswith('.mp4')]) if os.path.isdir(vid_dir) else []
    audio = sorted([f for f in os.listdir(aud_dir) if f.endswith('.mp3')]) if os.path.isdir(aud_dir) else []
    finals = sorted(glob_mod.glob(os.path.join(base, "final*.mp4")))
    return {"videos": videos, "audio": audio, "final": finals[-1] if finals else None}

def count_cut_versions(pat_key, cut_num):
    base = os.path.join(get_output_base(), pat_key, "videos")
    if not os.path.isdir(base):
        return 0
    prefix = f"カット{str(cut_num).zfill(2)}"
    return len([f for f in os.listdir(base) if f.startswith(prefix) and f.endswith('.mp4')])

def list_cut_versions(pat_key, cut_num):
    base = os.path.join(get_output_base(), pat_key, "videos")
    if not os.path.isdir(base):
        return []
    prefix = f"カット{str(cut_num).zfill(2)}"
    files = sorted([f for f in os.listdir(base) if f.startswith(prefix) and f.endswith('.mp4')])
    cid = get_active_cid()
    return [{"name": f, "url": f"/media/{cid}/{pat_key}/videos/{f}"} for f in files]

# ============================================================
# 修正ログ
# ============================================================
def log_correction(pat_key, cut_num, field, original, corrected, reason=""):
    cid = get_active_cid()
    if not cid:
        return
    path = _corrections_path(cid)
    entries = []
    if os.path.exists(path):
        with open(path, encoding="utf-8") as f:
            entries = json.load(f)
    entries.append({
        "pattern": pat_key, "cut": str(cut_num), "field": field,
        "original": original, "corrected": corrected, "reason": reason,
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    })
    with open(path, "w", encoding="utf-8") as f:
        json.dump(entries, f, ensure_ascii=False, indent=2)

# ============================================================
# ロゴマップ
# ============================================================
LOGO_MAP = {
    "バンタンデザイン研究所": "clients/vantan/2026/VDI/専門部/2025_VDI_PRO_logo_10801080.jpg",
    "バンタンゲームアカデミー": "clients/vantan/2026/VGA/専門部/2025_VGA_PRO_logo_10801080.jpg",
    "ヴィーナスアカデミー": "clients/vantan/2026/VA/専門部/2025_VA_PRO_logo_10801080.jpg",
    "バンタンクリエイターアカデミー": "clients/vantan/2026/VCA/専門部/2026_VCA_PRO_logo_10801080.jpg.jpg",
    "レコールバンタン": "clients/vantan/2026/LV/専門部/2023_LV_PRO_logo_10801080.jpg",
    "KADOKAWAアニメ・声優アカデミー": "clients/vantan/2026/KAA/専門部/2026_KAA_PRO_logo_10801080.jpg",
    "KADOKAWAマンガアカデミー": "clients/vantan/2026/KMA/専門部/2026_KMA_PRO_logo_10801080.jpg",
    "バンタンミュージックアカデミー": "clients/vantan/2026/VMA/専門部/2026_VMA_PRO_logo_10801080.jpg",
}

CINEMA_SUFFIX = (
    "Shot on a high-end cinema camera with anamorphic lens. "
    "Shallow depth of field, natural bokeh. "
    "Documentary film aesthetic, warm cinematic color grading. "
    "Subtle natural camera movement. No text, no logos, no signage."
)

# ============================================================
# 生成ロジック（バックグラウンド）
# ============================================================
def run_pipeline(pat_key, steps=None):
    from dotenv import load_dotenv
    import requests as req
    load_dotenv(os.path.join(ROOT, ".env"))
    fal_key = os.getenv("FAL_API_KEY")
    creatomate_key = os.getenv("CREATOMATE_API_KEY")

    if steps is None:
        steps = ["video", "narration", "compose"]

    pat = PATTERNS.get(pat_key)
    if not pat:
        set_pattern_state(pat_key, status="error", message="パターンが見つかりません")
        return

    cuts = pat["cuts"]
    out_dir = os.path.join(get_output_base(), pat_key)
    os.makedirs(os.path.join(out_dir, "videos"), exist_ok=True)
    os.makedirs(os.path.join(out_dir, "audio"), exist_ok=True)

    set_pattern_state(pat_key, status="running", progress="0/?", message="開始中...", stop_requested=False)

    def should_stop():
        s = load_state()
        return s.get(pat_key, {}).get("stop_requested", False)

    def upload(filepath, content_type):
        full = os.path.join(VANTAN, filepath) if not os.path.isabs(filepath) else filepath
        with open(full, 'rb') as f:
            d = f.read()
        try:
            init = req.post(
                "https://rest.alpha.fal.ai/storage/upload/initiate",
                headers={"Authorization": f"Key {fal_key}", "Content-Type": "application/json"},
                json={"file_name": os.path.basename(filepath), "content_type": content_type},
                timeout=30,
            )
            if init.status_code == 200:
                req.put(init.json()["upload_url"], data=d, headers={"Content-Type": content_type}, timeout=60)
                return init.json()["file_url"]
        except Exception:
            pass
        try:
            r = req.post("https://0x0.st", files={"file": (os.path.basename(filepath), d, content_type)}, timeout=60)
            if r.status_code == 200:
                return r.text.strip()
        except Exception:
            pass
        raise Exception(f"ファイルアップロード失敗: {os.path.basename(filepath)}")

    try:
        # Step 1: 動画生成（Veo3.1）
        if "video" in steps:
            from google import genai
            gemini_key = os.getenv("GEMINI_API_KEY")
            gclient = genai.Client(api_key=gemini_key)

            set_pattern_state(pat_key, message="Step 1/3: 動画生成（Veo3.1）")
            for i, cut in enumerate(cuts):
                if should_stop():
                    set_pattern_state(pat_key, status="stopped", message="停止")
                    return
                num = cut["num"]
                vid_path = os.path.join(out_dir, "videos", f"カット{num.zfill(2)}.mp4")
                if os.path.exists(vid_path):
                    set_pattern_state(pat_key, progress=f"動画 {i+1}/{len(cuts)}", message=f"カット{num} スキップ")
                    continue

                cut_st = get_cut_state(pat_key, num)
                prompt_en = cut_st.get("prompt_en_override") or cut.get("prompt_en", "")
                if not prompt_en:
                    continue
                prompt_en = re.sub(r'(?i)(the camera work is slightly shaky.*?\.)', '', prompt_en)
                prompt_en = re.sub(r'(?i)(the shot looks raw.*?\.)', '', prompt_en)
                prompt_en = re.sub(r'(?i)(natural lighting, authentic social media aesthetic\.?)', '', prompt_en)
                prompt_en = re.sub(r'(?i)(as if shot on an? iphone.*?\.)', '', prompt_en)
                full_prompt = f"{prompt_en.strip()} {CINEMA_SUFFIX}"

                set_pattern_state(pat_key, progress=f"動画 {i+1}/{len(cuts)}", message=f"カット{num} Veo3生成開始...")

                try:
                    op = gclient.models.generate_videos(
                        model="veo-3.1-generate-preview",
                        prompt=full_prompt,
                        config=genai.types.GenerateVideosConfig(aspect_ratio="9:16", number_of_videos=1),
                    )
                    for poll_i in range(90):
                        if should_stop():
                            set_pattern_state(pat_key, status="stopped", message="停止")
                            return
                        time.sleep(10)
                        op = gclient.operations.get(op)
                        set_pattern_state(pat_key, message=f"カット{num} Veo3: {'完了' if op.done else '生成中'} ({poll_i+1})")
                        if op.done:
                            if op.response and op.response.generated_videos:
                                video = op.response.generated_videos[0]
                                vid_data = gclient.files.download(file=video.video)
                                vid_bytes = b""
                                for chunk in vid_data:
                                    vid_bytes += chunk
                                with open(vid_path, 'wb') as f:
                                    f.write(vid_bytes)
                                set_pattern_state(pat_key, message=f"カット{num} OK ({len(vid_bytes)//1024}KB)")
                            else:
                                set_pattern_state(pat_key, message=f"カット{num} NG 動画データなし")
                            break
                    else:
                        set_pattern_state(pat_key, message=f"カット{num} NG タイムアウト")
                except Exception as e:
                    err_msg = str(e)[:200]
                    set_pattern_state(pat_key, status="error", message=f"カット{num} 動画エラー: {err_msg}")
                    if "quota" in err_msg.lower() or "429" in err_msg:
                        set_pattern_state(pat_key, status="error", message="APIクォータ超過")
                        return

        # Step 2: ナレーション生成（ElevenLabs）
        if "narration" in steps:
            set_pattern_state(pat_key, message="Step 2/3: ナレーション生成（ElevenLabs）")
            elevenlabs_key = os.getenv("ELEVENLABS_API_KEY")
            voice_id = "0ptCJp0xgdabdcpVtCB5"
            for i, cut in enumerate(cuts):
                if should_stop():
                    set_pattern_state(pat_key, status="stopped", message="停止")
                    return
                num = cut["num"]
                aud_path = os.path.join(out_dir, "audio", f"カット{num.zfill(2)}.mp3")
                if os.path.exists(aud_path):
                    set_pattern_state(pat_key, progress=f"音声 {i+1}/{len(cuts)}", message=f"カット{num} スキップ")
                    continue
                cut_st = get_cut_state(pat_key, num)
                narration = cut_st.get("narration_override") or cut.get("narration", "")
                if not narration:
                    continue
                set_pattern_state(pat_key, progress=f"音声 {i+1}/{len(cuts)}", message=f"カット{num} 音声生成中...")
                try:
                    resp = req.post(
                        f"https://api.elevenlabs.io/v1/text-to-speech/{voice_id}",
                        headers={"xi-api-key": elevenlabs_key, "Content-Type": "application/json"},
                        json={"text": narration, "model_id": "eleven_multilingual_v2",
                              "voice_settings": {"stability": 0.6, "similarity_boost": 0.75, "style": 0.15, "use_speaker_boost": True}},
                        timeout=60,
                    )
                    if resp.status_code == 200 and resp.headers.get("content-type", "").startswith("audio"):
                        with open(aud_path, 'wb') as f:
                            f.write(resp.content)
                        set_pattern_state(pat_key, message=f"カット{num} OK ({len(resp.content)//1024}KB)")
                    else:
                        set_pattern_state(pat_key, status="error", message=f"カット{num} ElevenLabs {resp.status_code}")
                        if resp.status_code in (401, 403):
                            return
                except Exception as e:
                    set_pattern_state(pat_key, message=f"カット{num} 音声エラー: {str(e)[:100]}")

        # Step 3: Creatomate 合成
        if "compose" in steps:
            missing_vid, missing_aud = [], []
            for cut in cuts:
                num = cut["num"]
                vp = os.path.join(out_dir, "videos", f"カット{num.zfill(2)}.mp4")
                ap = os.path.join(out_dir, "audio", f"カット{num.zfill(2)}.mp3")
                if not os.path.exists(vp): missing_vid.append(num)
                if not os.path.exists(ap): missing_aud.append(num)
            if missing_vid or missing_aud:
                msg_parts = []
                if missing_vid: msg_parts.append(f"動画なし: カット{','.join(missing_vid)}")
                if missing_aud: msg_parts.append(f"音声なし: カット{','.join(missing_aud)}")
                set_pattern_state(pat_key, status="error", message=f"素材不足: {'; '.join(msg_parts)}")
                return

            set_pattern_state(pat_key, message="Step 3/3: 合成（Creatomate）")
            vid_urls, audio_urls = {}, {}
            for i, cut in enumerate(cuts):
                if should_stop():
                    set_pattern_state(pat_key, status="stopped", message="停止")
                    return
                num = cut["num"]
                vid_path = os.path.join(out_dir, "videos", f"カット{num.zfill(2)}.mp4")
                aud_path = os.path.join(out_dir, "audio", f"カット{num.zfill(2)}.mp3")
                if os.path.exists(vid_path): vid_urls[num] = upload(vid_path, "video/mp4")
                if os.path.exists(aud_path): audio_urls[num] = upload(aud_path, "audio/mpeg")
                set_pattern_state(pat_key, progress=f"アップロード {i+1}/{len(cuts)}", message=f"カット{num}")

            logo_url = ""
            logo_rel = LOGO_MAP.get(pat["school"], "")
            if logo_rel:
                logo_full = os.path.join(ROOT, logo_rel)
                if os.path.exists(logo_full):
                    logo_url = upload(logo_full, "image/jpeg")

            se_base = os.path.join(ROOT, "clients/vantan/se/真面目バージョン")
            se_files = {}
            for cat in ['01_impact', '02_negative', '03_neutral', '04_tiktok']:
                cat_dir = os.path.join(se_base, cat)
                if os.path.isdir(cat_dir):
                    se_files[cat] = sorted([f for f in os.listdir(cat_dir) if f.endswith('.mp3')])

            CUT_SE_MAP = {
                '1': '04_tiktok', '2': '02_negative', '3': '02_negative',
                '4': '01_impact', '5': '01_impact', '6': '03_neutral',
                '7': '03_neutral', '8': '03_neutral', '9': '01_impact',
                '10': '01_impact', '11': '04_tiktok',
            }
            se_urls, prev_file = [], None
            for cut in cuts:
                cat = CUT_SE_MAP.get(cut['num'], '03_neutral')
                candidates = se_files.get(cat, [])
                available = [f for f in candidates if f != prev_file] or candidates
                chosen = random.choice(available) if available else None
                se_urls.append(upload(os.path.join(se_base, cat, chosen), "audio/mpeg") if chosen else "")
                prev_file = chosen

            bgm_path = os.path.join(ROOT, "clients/vantan/bgm/01_hopeful/bgm01_piano_Cmaj_70bpm.mp3")
            bgm_url = upload(bgm_path, "audio/mpeg") if os.path.exists(bgm_path) else ""

            set_pattern_state(pat_key, message="Creatomate レンダリング中...")

            elements = []
            for i, cut in enumerate(cuts):
                num = cut['num']
                if num not in vid_urls or num not in audio_urls:
                    continue
                cut_st = get_cut_state(pat_key, num)
                telop_text = cut_st.get("telop_override") or cut['telop']
                scene_elements = [
                    {"type": "video", "source": vid_urls[num], "fit": "cover", "duration": "100%", "loop": True},
                    {"type": "audio", "source": audio_urls[num], "volume": "100%"},
                ]
                if cut['logo'] == '○':
                    if logo_url:
                        scene_elements.append({
                            "type": "image", "source": logo_url,
                            "x": "50%", "y": "50%", "width": "75%", "height": "25%",
                            "fit": "contain", "x_alignment": "50%", "y_alignment": "50%", "z_index": 15,
                        })
                    else:
                        scene_elements.append({
                            "type": "text", "text": pat["school"],
                            "width": "85%", "height": "25%", "x": "50%", "y": "50%",
                            "duration": "100%", "z_index": 15,
                            "fill_color": "#FFFFFF", "font_family": "Noto Sans JP", "font_weight": "900",
                            "shadow_color": "rgba(0,0,0,0.7)", "shadow_blur": "30px",
                            "x_alignment": "50%", "y_alignment": "50%", "content_alignment": "center",
                            "dynamic_font_size": True, "font_size_maximum": "80px", "font_size_minimum": "36px",
                            "fit": "shrink",
                        })
                elif telop_text:
                    scene_elements.append({
                        "type": "text", "text": telop_text,
                        "width": "85%", "height": "20%", "x": "50%", "y": "50%",
                        "duration": "100%", "z_index": 15,
                        "fill_color": "#FFFFFF", "font_family": "Noto Sans JP", "font_weight": "900",
                        "shadow_color": "rgba(0,0,0,0.6)", "shadow_blur": "25px",
                        "x_alignment": "50%", "y_alignment": "50%", "content_alignment": "center",
                        "dynamic_font_size": True, "font_size_maximum": "70px", "font_size_minimum": "30px",
                        "fit": "shrink",
                    })
                if i < len(se_urls) and se_urls[i]:
                    scene_elements.append({"type": "audio", "source": se_urls[i], "volume": "30%", "duration": "100%"})
                scene = {"type": "composition", "track": 1, "elements": scene_elements}
                if len(elements) > 0:
                    scene["transition"] = {"type": "crossfade", "duration": 0.1}
                elements.append(scene)

            if bgm_url:
                elements.append({"type": "audio", "source": bgm_url, "track": 2, "volume": "30%", "duration": "100%"})

            if not elements:
                set_pattern_state(pat_key, status="error", message="合成する素材がありません")
                return

            cr_resp = req.post(
                "https://api.creatomate.com/v1/renders",
                headers={"Authorization": f"Bearer {creatomate_key}", "Content-Type": "application/json"},
                json={"source": {"output_format": "mp4", "frame_rate": 30, "width": 720, "height": 1280, "elements": elements}},
                timeout=60,
            )
            renders = cr_resp.json()
            render_obj = renders[0] if isinstance(renders, list) else renders
            render_id = render_obj.get("id", "")

            for poll_i in range(60):
                if should_stop():
                    set_pattern_state(pat_key, status="stopped", message="停止")
                    return
                time.sleep(10)
                poll = req.get(f"https://api.creatomate.com/v1/renders/{render_id}",
                              headers={"Authorization": f"Bearer {creatomate_key}"}, timeout=30)
                status = poll.json().get("status", "")
                set_pattern_state(pat_key, message=f"レンダリング: {status} ({poll_i+1}/60)")
                if status == "succeeded":
                    final_url = poll.json().get("url", "")
                    vid = req.get(final_url, timeout=120)
                    final_path = os.path.join(out_dir, "final.mp4")
                    with open(final_path, 'wb') as f:
                        f.write(vid.content)
                    set_pattern_state(pat_key, status="done", progress=f"{len(cuts)}/{len(cuts)}",
                                     message=f"完成 ({len(vid.content)//1024}KB)")
                    return
                elif status == "failed":
                    set_pattern_state(pat_key, status="error", message=f"失敗: {poll.json().get('error_message','')}")
                    return

            set_pattern_state(pat_key, status="error", message="タイムアウト（10分）")
            return

        if "compose" not in steps:
            set_pattern_state(pat_key, status="done", message="素材生成完了（合成はまだ）")

    except Exception as e:
        set_pattern_state(pat_key, status="error", message=str(e)[:200])

# ============================================================
# STATUS.md パーサー（CCM Dashboard 用）
# ============================================================
def parse_status_md() -> dict:
    if not STATUS_MD.exists():
        return {"error": "STATUS.md not found", "patterns": [], "summary": {}, "next_tasks": []}
    text = STATUS_MD.read_text(encoding="utf-8")
    summary = {}
    summary_match = re.search(r"## クイックサマリ\s*\n\|[^\n]+\n\|[-\s|]+\n((?:\|[^\n]+\n)*)", text)
    if summary_match:
        for line in summary_match.group(1).strip().split("\n"):
            cols = [c.strip().strip("*") for c in line.split("|")[1:-1]]
            if len(cols) >= 2:
                summary[cols[0]] = cols[1]
    patterns = []
    pattern_match = re.search(r"## パターン別進捗\s*\n\|[^\n]+\n\|[-\s|]+\n((?:\|[^\n]+\n)*)", text)
    if pattern_match:
        for line in pattern_match.group(1).strip().split("\n"):
            cols = [c.strip() for c in line.split("|")[1:-1]]
            if len(cols) >= 6:
                vid_parts = cols[1].split("/")
                aud_parts = cols[2].split("/")
                vid_done = int(vid_parts[0]) if vid_parts[0].isdigit() else 0
                vid_total = int(vid_parts[1]) if len(vid_parts) > 1 and vid_parts[1].isdigit() else 11
                aud_done = int(aud_parts[0]) if aud_parts[0].isdigit() else 0
                aud_total = int(aud_parts[1]) if len(aud_parts) > 1 and aud_parts[1].isdigit() else 11
                has_final = "✅" in cols[3]
                status = cols[4].strip("* ")
                patterns.append({
                    "id": cols[0].strip(), "video": f"{vid_done}/{vid_total}",
                    "video_done": vid_done, "video_total": vid_total,
                    "audio": f"{aud_done}/{aud_total}", "audio_done": aud_done, "audio_total": aud_total,
                    "has_final": has_final, "status": status, "memo": cols[5] if len(cols) > 5 else "",
                })
    total = len(patterns)
    completed = sum(1 for p in patterns if p["has_final"])
    in_progress = sum(1 for p in patterns if not p["has_final"] and (p["video_done"] > 0 or p["audio_done"] > 0))
    not_started = sum(1 for p in patterns if p["video_done"] == 0 and p["audio_done"] == 0 and not p["has_final"])
    next_tasks = []
    task_match = re.search(r"## 次にやるべきこと[^\n]*\n((?:\d+\.[^\n]+\n)*)", text)
    if task_match:
        for line in task_match.group(1).strip().split("\n"):
            task_text = re.sub(r"^\d+\.\s*", "", line).strip()
            if task_text:
                next_tasks.append(task_text)
    return {
        "summary": summary, "patterns": patterns, "total": total,
        "completed": completed, "in_progress": in_progress, "not_started": not_started, "next_tasks": next_tasks,
    }

# ============================================================
# FastAPI App
# ============================================================
app = FastAPI(title="kage-lab Unified Studio", version=VERSION)
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

# ── Auth ────────────────────────────────────────────────
def get_session(request: Request) -> str | None:
    return request.cookies.get("session")

def is_authed(request: Request) -> bool:
    return get_session(request) in valid_sessions

@app.get("/login", response_class=HTMLResponse)
async def login_page(error: str = ""):
    err_html = '<div style="color:#f85149;font-size:13px;margin-bottom:8px">Wrong password</div>' if error else ""
    return f"""<!DOCTYPE html><html lang="ja"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>kage-lab Studio</title>
<style>
body{{font-family:-apple-system,sans-serif;background:#0d1117;color:#e0e0e0;
display:flex;justify-content:center;align-items:center;height:100dvh;margin:0}}
.box{{background:#161b22;padding:32px;border-radius:16px;width:300px;text-align:center;border:1px solid #30363d}}
h2{{color:#b8956a;margin-bottom:20px;font-size:18px}}
.ver{{font-size:0.6em;opacity:0.5}}
input{{width:100%;padding:12px;margin-bottom:12px;border-radius:8px;border:1px solid #30363d;background:#0d1117;color:#e0e0e0;font-size:16px;box-sizing:border-box}}
button{{width:100%;padding:12px;border-radius:8px;border:none;background:#b8956a;color:#0d1117;font-size:16px;font-weight:bold;cursor:pointer}}
</style></head><body>
<div class="box"><h2>kage-lab Studio <span class="ver">v{VERSION}</span></h2>
<form method="POST" action="/login"><input type="password" name="password" placeholder="Password" autofocus>
{err_html}<button type="submit">Login</button></form></div></body></html>"""

@app.post("/login")
async def login(request: Request, password: str = Form(...)):
    if password == PASSWORD:
        sid = secrets.token_hex(16)
        valid_sessions.add(sid)
        resp = RedirectResponse("/", status_code=303)
        resp.set_cookie("session", sid, httponly=True, path="/")
        return resp
    return RedirectResponse("/login?error=1", status_code=303)

# ── Main UI ─────────────────────────────────────────────
@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    if not is_authed(request):
        return RedirectResponse("/login")
    html_path = os.path.join(ROOT, "unified_ui.html")
    with open(html_path, encoding="utf-8") as f:
        content = f.read()
    return Response(content=content, media_type="text/html", headers={
        "Cache-Control": "no-cache, no-store, must-revalidate",
    })

# ── CCM API (Dashboard / Chat / Actions) ───────────────
@app.get("/api/dashboard")
async def api_dashboard(request: Request):
    if not is_authed(request):
        return JSONResponse({"error": "not authenticated"}, status_code=401)
    data = parse_status_md()
    elapsed = int(time.time() - server_start)
    h, m = divmod(elapsed // 60, 60)
    data["uptime"] = f"{h}h{m:02d}m" if h else f"{m}m"
    data["claude_busy"] = claude_busy
    data["sessions"] = len(claude_sessions)
    data["tunnel_url"] = tunnel_url
    output_dir = WORK_DIR / "apps" / "vantan-video" / "output"
    try:
        outputs = sorted(output_dir.glob("*"), key=lambda p: p.stat().st_mtime, reverse=True)
        data["last_output"] = outputs[0].name if outputs else ""
    except Exception:
        data["last_output"] = ""
    return JSONResponse(data)

@app.get("/api/server-status")
async def api_server_status(request: Request):
    if not is_authed(request):
        return JSONResponse({"error": "not authenticated"}, status_code=401)
    elapsed = int(time.time() - server_start)
    h, m = divmod(elapsed // 60, 60)
    busy = False
    try:
        result = subprocess.run(["pgrep", "-f", "claude.*-p"], capture_output=True, text=True, timeout=3)
        busy = bool(result.stdout.strip())
    except Exception:
        pass
    return JSONResponse({
        "claude_busy": busy, "uptime": f"{h}h{m:02d}m" if h else f"{m}m",
        "active_sessions": len(claude_sessions), "tunnel_url": tunnel_url,
    })

@app.post("/api/ask")
async def api_ask(request: Request):
    if not is_authed(request):
        return JSONResponse({"error": "not authenticated"}, status_code=401)
    body = await request.json()
    message = body.get("message", "")
    web_session = get_session(request)

    async def stream():
        global claude_busy
        claude_busy = True
        try:
            claude_cmd = find_claude()
            env = os.environ.copy()
            env["PATH"] = "/opt/homebrew/bin:" + env.get("PATH", "")
            cmd = [claude_cmd, "-p", message, "--output-format", "stream-json", "--verbose", "--include-partial-messages"]
            claude_sid = claude_sessions.get(web_session)
            if claude_sid:
                cmd.extend(["--resume", claude_sid])
            proc = await asyncio.create_subprocess_exec(
                *cmd, stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.DEVNULL,
                cwd=str(WORK_DIR), env=env,
            )
            full_text = ""
            session_id = ""
            async def read_with_timeout():
                try:
                    return await asyncio.wait_for(proc.stdout.readline(), timeout=300)
                except asyncio.TimeoutError:
                    proc.kill()
                    return None
            while True:
                line = await read_with_timeout()
                if line is None:
                    yield f"data: {json.dumps({'type': 'error', 'text': 'Timeout (5 min)'})}\n\n"
                    break
                if not line:
                    break
                line_str = line.decode("utf-8", errors="replace").strip()
                if not line_str:
                    continue
                try:
                    event = json.loads(line_str)
                except json.JSONDecodeError:
                    continue
                event_type = event.get("type", "")
                if event_type == "system" and event.get("session_id"):
                    session_id = event["session_id"]
                if event_type == "stream_event":
                    se = event.get("event", {})
                    if se.get("type") == "content_block_delta":
                        delta_obj = se.get("delta", {})
                        if delta_obj.get("type") == "text_delta":
                            chunk = delta_obj.get("text", "")
                            if chunk:
                                full_text += chunk
                                yield f"data: {json.dumps({'type': 'delta', 'text': chunk})}\n\n"
                if event_type == "result":
                    result_text = event.get("result", "")
                    if result_text and len(result_text) > len(full_text):
                        delta = result_text[len(full_text):]
                        full_text = result_text
                        yield f"data: {json.dumps({'type': 'delta', 'text': delta})}\n\n"
                    if event.get("session_id"):
                        session_id = event["session_id"]
            await proc.wait()
            if session_id and web_session:
                claude_sessions[web_session] = session_id
            if not full_text:
                yield f"data: {json.dumps({'type': 'delta', 'text': '(no response)'})}\n\n"
            yield f"data: {json.dumps({'type': 'done', 'session_id': session_id})}\n\n"
        except Exception as e:
            yield f"data: {json.dumps({'type': 'error', 'text': str(e)})}\n\n"
        finally:
            claude_busy = False

    return StreamingResponse(stream(), media_type="text/event-stream")

@app.post("/api/queue")
async def api_queue_add(request: Request):
    if not is_authed(request):
        return JSONResponse({"error": "not authenticated"}, status_code=401)
    body = await request.json()
    text = body.get("message", "").strip()
    if not text:
        return JSONResponse({"error": "empty"}, status_code=400)
    ws = get_session(request) or ""
    if ws not in message_queue:
        message_queue[ws] = []
    message_queue[ws].append({"text": text, "timestamp": time.time()})
    return JSONResponse({"ok": True, "queued": len(message_queue[ws])})

@app.get("/api/queue")
async def api_queue_get(request: Request):
    if not is_authed(request):
        return JSONResponse({"error": "not authenticated"}, status_code=401)
    ws = get_session(request) or ""
    items = message_queue.get(ws, [])
    return JSONResponse({"items": items, "count": len(items)})

@app.delete("/api/queue")
async def api_queue_clear(request: Request):
    if not is_authed(request):
        return JSONResponse({"error": "not authenticated"}, status_code=401)
    ws = get_session(request) or ""
    message_queue.pop(ws, None)
    return JSONResponse({"ok": True})

@app.post("/api/new-chat")
async def api_new_chat(request: Request):
    if not is_authed(request):
        return JSONResponse({"error": "not authenticated"}, status_code=401)
    ws = get_session(request)
    claude_sessions.pop(ws, None)
    message_queue.pop(ws, None)
    return JSONResponse({"ok": True})

@app.post("/api/quick-action")
async def api_quick_action(request: Request):
    if not is_authed(request):
        return JSONResponse({"error": "not authenticated"}, status_code=401)
    body = await request.json()
    action = body.get("action", "")
    try:
        if action == "git-status":
            r = subprocess.run(["git", "status", "-s"], capture_output=True, text=True, cwd=str(WORK_DIR), timeout=10)
            return JSONResponse({"result": r.stdout or "(clean)"})
        elif action == "git-pull":
            r = subprocess.run(["git", "pull"], capture_output=True, text=True, cwd=str(WORK_DIR), timeout=30)
            return JSONResponse({"result": r.stdout + r.stderr})
        elif action == "git-push":
            r = subprocess.run(["git", "push"], capture_output=True, text=True, cwd=str(WORK_DIR), timeout=30)
            return JSONResponse({"result": r.stdout + r.stderr})
        elif action == "git-log":
            r = subprocess.run(["git", "log", "--oneline", "-10"], capture_output=True, text=True, cwd=str(WORK_DIR), timeout=10)
            return JSONResponse({"result": r.stdout})
        elif action == "refresh-status":
            return JSONResponse({"result": "ok", "data": parse_status_md()})
        else:
            return JSONResponse({"error": f"Unknown: {action}"}, status_code=400)
    except subprocess.TimeoutExpired:
        return JSONResponse({"result": "(timeout)"})
    except Exception as e:
        return JSONResponse({"result": f"(error: {e})"})

@app.post("/api/launch")
async def api_launch(request: Request):
    if not is_authed(request):
        return JSONResponse({"error": "not authenticated"}, status_code=401)
    body = await request.json()
    service_id = body.get("service", "")
    svc = SERVICES.get(service_id)
    if not svc:
        return JSONResponse({"error": f"Unknown: {service_id}"}, status_code=400)
    try:
        result = subprocess.run(["lsof", "-ti", f":{svc['port']}"], capture_output=True, text=True, timeout=3)
        if result.stdout.strip():
            return JSONResponse({"ok": True, "status": "already_running", "name": svc["name"], "port": svc["port"]})
    except Exception:
        pass
    try:
        env = os.environ.copy()
        env["PATH"] = "/opt/homebrew/bin:" + env.get("PATH", "")
        proc = subprocess.Popen(svc["cmd"], cwd=svc["cwd"], env=env, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        launched_procs[service_id] = proc
        return JSONResponse({"ok": True, "status": "started", "name": svc["name"], "port": svc["port"], "pid": proc.pid})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

@app.get("/api/services")
async def api_services():
    result = {}
    for sid, svc in SERVICES.items():
        running = False
        try:
            r = subprocess.run(["lsof", "-ti", f":{svc['port']}"], capture_output=True, text=True, timeout=3)
            running = bool(r.stdout.strip())
        except Exception:
            pass
        result[sid] = {"name": svc["name"], "port": svc["port"], "running": running}
    return JSONResponse(result)

# ── Studio API (Video Generation) ──────────────────────
@app.get("/api/cartridges")
def api_cartridges():
    reg = load_registry()
    meta = load_meta()
    files = []
    for f in sorted(os.listdir(BRIEFS_DIR)):
        if not f.endswith(".xlsx"):
            continue
        fpath = os.path.join(BRIEFS_DIR, f)
        stat = os.stat(fpath)
        cid = ""
        for c, info in reg.get("cartridges", {}).items():
            if info["filename"] == f:
                cid = c
                break
        files.append({"filename": f, "size_kb": round(stat.st_size / 1024, 1),
                       "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"), "cid": cid})
    return {"files": files, "meta": meta, "active_cid": meta.get("cartridge_id", "")}

@app.post("/api/cartridges/upload")
async def api_upload(file: UploadFile = File(...)):
    if not file.filename.endswith(".xlsx"):
        return {"error": ".xlsx ファイルのみ"}
    dest = os.path.join(BRIEFS_DIR, file.filename)
    content = await file.read()
    with open(dest, "wb") as f:
        f.write(content)
    return {"ok": True, "filename": file.filename, "size_kb": round(len(content) / 1024, 1)}

@app.post("/api/cartridges/load")
def api_load_cartridge(data: dict):
    filename = data.get("filename", "")
    fpath = os.path.join(BRIEFS_DIR, filename)
    if not os.path.exists(fpath):
        return {"error": f"Not found: {filename}"}
    try:
        patterns = parse_xlsx_to_patterns(fpath)
        if not patterns:
            return {"error": "パターンなし"}
        cid = register_cartridge(filename)
        save_cartridge_patterns(cid, patterns)
        output_dest = cartridge_output_base(cid)
        legacy = os.path.join(VANTAN, "output", "workflow_002")
        if os.path.isdir(legacy):
            for pat_key in os.listdir(legacy):
                src = os.path.join(legacy, pat_key)
                dst = os.path.join(output_dest, pat_key)
                if os.path.isdir(src) and not os.path.exists(dst):
                    os.symlink(src, dst)
        sf = _state_path(cid)
        if not os.path.exists(sf):
            old_state_file = os.path.join(ROOT, "control_panel_state.json")
            if os.path.exists(old_state_file):
                shutil.copy2(old_state_file, sf)
        total_cuts = sum(len(p.get("cuts", [])) for p in patterns.values())
        set_active_cid(cid, filename, len(patterns))
        reload_active_patterns()
        schools = sorted(set(p["school"] for p in patterns.values() if p.get("school")))
        return {"ok": True, "cid": cid, "filename": filename,
                "patterns": len(patterns), "total_cuts": total_cuts, "schools": schools}
    except Exception as e:
        return {"error": str(e)}

@app.post("/api/cartridges/activate/{cid}")
def api_activate_cartridge(cid: str):
    reg = load_registry()
    if cid not in reg.get("cartridges", {}):
        return {"error": f"Not found: {cid}"}
    filename = reg["cartridges"][cid]["filename"]
    patterns = load_cartridge_patterns(cid)
    set_active_cid(cid, filename, len(patterns))
    reload_active_patterns()
    return {"ok": True, "cid": cid, "patterns": len(PATTERNS)}

@app.post("/api/cartridges/refresh")
def api_refresh_cartridge():
    meta = load_meta()
    cid = meta.get("cartridge_id", "")
    filename = meta.get("loaded_file", "")
    if not cid or not filename:
        return {"error": "カートリッジなし"}
    fpath = os.path.join(BRIEFS_DIR, filename)
    if not os.path.exists(fpath):
        return {"error": f"Not found: {filename}"}
    try:
        patterns = parse_xlsx_to_patterns(fpath)
        save_cartridge_patterns(cid, patterns)
        if get_active_cid() == cid:
            reload_active_patterns()
        return {"ok": True, "cid": cid, "patterns": len(patterns)}
    except Exception as e:
        return {"error": str(e)}

@app.get("/api/status")
def api_status():
    state = load_state()
    cid = get_active_cid()
    result = {}
    for pat_key, pat in PATTERNS.items():
        outputs = check_outputs(pat_key)
        total = len(pat["cuts"])
        pat_state = state.get(pat_key, {"status": "idle", "progress": "", "message": ""})
        cuts_qc = pat_state.get("cuts_qc", {})
        cuts_data = []
        for c in pat["cuts"]:
            cnum = c["num"]
            cqc = cuts_qc.get(str(cnum), {})
            vid_name = f"カット{cnum.zfill(2)}.mp4"
            aud_name = f"カット{cnum.zfill(2)}.mp3"
            cuts_data.append({
                **c, "has_video": vid_name in outputs["videos"], "has_audio": aud_name in outputs["audio"],
                "qc_status": cqc.get("qc_status", "pending"),
                "narration_override": cqc.get("narration_override"), "telop_override": cqc.get("telop_override"),
                "prompt_jp_override": cqc.get("prompt_jp_override"), "prompt_en_override": cqc.get("prompt_en_override"),
                "video_versions": count_cut_versions(pat_key, cnum),
            })
        qc_counts = {"pending": 0, "approved": 0, "rejected": 0}
        for cd in cuts_data:
            qc_counts[cd["qc_status"]] = qc_counts.get(cd["qc_status"], 0) + 1
        result[pat_key] = {
            "school": pat["school"], "field": pat["field"], "child": pat["child"],
            "total_cuts": total, "videos": len(outputs["videos"]),
            "audio": len(outputs["audio"]), "has_final": outputs["final"] is not None,
            "cuts": cuts_data, "qc_counts": qc_counts,
            **{k: v for k, v in pat_state.items() if k != "cuts_qc"},
        }
    return {"patterns": result, "active_cid": cid}

@app.post("/api/generate/{pat_key}")
def api_generate(pat_key: str, data: dict = None):
    ps = get_pattern_state(pat_key)
    if ps.get("status") == "running":
        return {"error": "既に実行中です"}
    steps = data.get("steps") if data else None
    set_pattern_state(pat_key, status="running", progress="0/?", message="開始中...", stop_requested=False)
    t = threading.Thread(target=run_pipeline, args=(pat_key, steps), daemon=True)
    t.start()
    return {"ok": True, "pattern": pat_key, "steps": steps}

@app.post("/api/stop/{pat_key}")
def api_stop(pat_key: str):
    set_pattern_state(pat_key, stop_requested=True, message="停止リクエスト中...")
    return {"ok": True}

@app.post("/api/regenerate/{pat_key}")
def api_regenerate(pat_key: str, data: dict = None):
    ps = get_pattern_state(pat_key)
    if ps.get("status") == "running":
        return {"error": "既に実行中です"}
    target = (data or {}).get("target", "all")
    out_dir = os.path.join(get_output_base(), pat_key)
    if target in ("video", "all"):
        d = os.path.join(out_dir, "videos")
        if os.path.isdir(d):
            shutil.rmtree(d)
    if target in ("audio", "all"):
        d = os.path.join(out_dir, "audio")
        if os.path.isdir(d):
            shutil.rmtree(d)
    final = os.path.join(out_dir, "final.mp4")
    if os.path.exists(final):
        os.remove(final)
    steps_map = {"video": ["video"], "audio": ["narration"], "all": ["video", "narration"]}
    steps = steps_map.get(target, ["video", "narration"])
    set_pattern_state(pat_key, status="running", progress="0/?", message="再生成中...", stop_requested=False)
    t = threading.Thread(target=run_pipeline, args=(pat_key, steps), daemon=True)
    t.start()
    return {"ok": True, "pattern": pat_key, "target": target}

@app.post("/api/generate-batch")
def api_generate_batch(data: dict):
    pat_keys = data.get("patterns", [])
    steps = data.get("steps")
    def run_batch():
        for pk in pat_keys:
            s = load_state()
            if s.get(pk, {}).get("stop_requested"):
                break
            run_pipeline(pk, steps)
    for pk in pat_keys:
        set_pattern_state(pk, status="queued", message="バッチ待機中...", stop_requested=False)
    if pat_keys:
        set_pattern_state(pat_keys[0], status="running", message="バッチ開始...")
    t = threading.Thread(target=run_batch, daemon=True)
    t.start()
    return {"ok": True, "patterns": pat_keys}

@app.post("/api/regenerate-cut/{pat_key}/{cut_num}")
def api_regenerate_cut(pat_key: str, cut_num: str):
    ps = get_pattern_state(pat_key)
    if ps.get("status") == "running":
        return {"error": "既に実行中です"}
    out_dir = os.path.join(get_output_base(), pat_key)
    for ext, sub in [(".mp4", "videos"), (".mp3", "audio")]:
        path = os.path.join(out_dir, sub, f"カット{cut_num.zfill(2)}{ext}")
        if os.path.exists(path):
            ver = count_cut_versions(pat_key, cut_num)
            archive_name = f"カット{cut_num.zfill(2)}_v{str(ver).zfill(3)}{ext}"
            os.rename(path, os.path.join(out_dir, sub, archive_name))
    final = os.path.join(out_dir, "final.mp4")
    if os.path.exists(final):
        os.remove(final)
    set_pattern_state(pat_key, status="running", progress=f"カット{cut_num}再生成", message=f"カット{cut_num}再生成中...", stop_requested=False)
    t = threading.Thread(target=run_pipeline, args=(pat_key, ["video", "narration"]), daemon=True)
    t.start()
    return {"ok": True, "pattern": pat_key, "cut": cut_num}

@app.post("/api/reload-sheets")
def api_reload_sheets():
    return api_refresh_cartridge()

@app.post("/api/stop-all")
def api_stop_all():
    state = load_state()
    for k in state:
        state[k]["stop_requested"] = True
    save_state(state)
    return {"ok": True}

@app.put("/api/cut/{pat_key}/{cut_num}/status")
def api_cut_status(pat_key: str, cut_num: str, data: dict):
    new_status = data.get("qc_status", "pending")
    if new_status not in ("pending", "approved", "rejected"):
        return {"error": "Invalid status"}
    set_cut_state(pat_key, cut_num, qc_status=new_status)
    return {"ok": True}

@app.put("/api/cut/{pat_key}/{cut_num}/text")
def api_cut_text(pat_key: str, cut_num: str, data: dict):
    pat = PATTERNS.get(pat_key)
    if not pat:
        return {"error": "Pattern not found"}
    cut = next((c for c in pat["cuts"] if c["num"] == cut_num), None)
    if not cut:
        return {"error": "Cut not found"}
    reason = data.get("reason", "")
    updates = {}
    for field in ["narration", "telop", "prompt_jp", "prompt_en"]:
        if field in data and data[field] is not None:
            override_key = f"{field}_override"
            original = cut.get(field, "")
            new_val = data[field]
            if new_val != original:
                updates[override_key] = new_val
                log_correction(pat_key, cut_num, field, original, new_val, reason)
            else:
                updates[override_key] = None
    if updates:
        set_cut_state(pat_key, cut_num, **updates)
    return {"ok": True}

@app.get("/api/cut/{pat_key}/{cut_num}/versions")
def api_cut_versions(pat_key: str, cut_num: str):
    videos = list_cut_versions(pat_key, cut_num)
    cid = get_active_cid()
    base = os.path.join(get_output_base(), pat_key, "audio")
    audio = []
    if os.path.isdir(base):
        prefix = f"カット{cut_num.zfill(2)}"
        files = sorted([f for f in os.listdir(base) if f.startswith(prefix) and f.endswith('.mp3')])
        audio = [{"name": f, "url": f"/media/{cid}/{pat_key}/audio/{f}"} for f in files]
    return {"videos": videos, "audio": audio}

@app.get("/api/corrections/{pat_key}")
def api_corrections(pat_key: str):
    cid = get_active_cid()
    if not cid:
        return []
    path = _corrections_path(cid)
    if os.path.exists(path):
        with open(path, encoding="utf-8") as f:
            entries = json.load(f)
        return [e for e in entries if e.get("pattern") == pat_key]
    return []

@app.put("/api/pattern/{pat_key}/approve-all")
def api_approve_all(pat_key: str):
    pat = PATTERNS.get(pat_key)
    if not pat:
        return {"error": "Pattern not found"}
    for c in pat["cuts"]:
        set_cut_state(pat_key, c["num"], qc_status="approved")
    return {"ok": True}

@app.put("/api/pattern/{pat_key}/pending-all")
def api_pending_all(pat_key: str):
    pat = PATTERNS.get(pat_key)
    if not pat:
        return {"error": "Pattern not found"}
    for c in pat["cuts"]:
        set_cut_state(pat_key, c["num"], qc_status="pending")
    return {"ok": True}

@app.get("/api/files/{pat_key}")
def api_files(pat_key: str):
    cid = get_active_cid()
    base = os.path.join(get_output_base(), pat_key)
    result = {"videos": [], "audio": [], "final": None}
    vid_dir, aud_dir = os.path.join(base, "videos"), os.path.join(base, "audio")
    if os.path.isdir(vid_dir):
        for f in sorted(os.listdir(vid_dir)):
            if f.endswith('.mp4'):
                result["videos"].append({"name": f, "url": f"/media/{cid}/{pat_key}/videos/{f}"})
    if os.path.isdir(aud_dir):
        for f in sorted(os.listdir(aud_dir)):
            if f.endswith('.mp3'):
                result["audio"].append({"name": f, "url": f"/media/{cid}/{pat_key}/audio/{f}"})
    final = os.path.join(base, "final.mp4")
    if os.path.exists(final):
        result["final"] = f"/media/{cid}/{pat_key}/final.mp4"
    return result

# ── Static files ────────────────────────────────────────
output_dir = os.path.join(VANTAN, "output")
if os.path.isdir(output_dir):
    app.mount("/media", StaticFiles(directory=output_dir, follow_symlink=True), name="media")

assets_dir = os.path.join(ROOT, "assets")
if os.path.isdir(assets_dir):
    app.mount("/assets", StaticFiles(directory=assets_dir), name="assets")

# ============================================================
# Main
# ============================================================
def main():
    print(f"\n  kage-lab Unified Studio v{VERSION}")
    print(f"  ─────────────────────────────────────")
    print(f"  Local:    http://localhost:{PORT}")

    local_ip = ""
    try:
        r = subprocess.run(["ipconfig", "getifaddr", "en0"], capture_output=True, text=True, timeout=3)
        local_ip = r.stdout.strip()
        if local_ip:
            print(f"  LAN:      http://{local_ip}:{PORT}")
    except Exception:
        pass

    print(f"  Password: {PASSWORD}")
    print()

    # Cloudflare Tunnel (background)
    tunnel_thread = threading.Thread(target=start_tunnel, daemon=True)
    tunnel_thread.start()

    print(f"  Tunnel starting...")
    print()

    uvicorn.run(app, host="0.0.0.0", port=PORT, log_level="warning")

if __name__ == "__main__":
    main()
